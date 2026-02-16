from flask import Flask, jsonify, render_template, request, redirect, url_for, flash, make_response
import pandas as pd
from io import BytesIO
from flask import send_file 
import mysql.connector
from datetime import date, datetime, timedelta
import os
import string
from flask_sqlalchemy import SQLAlchemy
import random
from dotenv import load_dotenv
from flask_mail import Mail, Message
from functools import wraps
from xhtml2pdf import pisa
from num2words import num2words
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from calendar import month_name
import csv
from io import StringIO, BytesIO

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('APP_SECRET_KEY')  # Secret key for flash messages

db_config = {
    'host': os.getenv('DB_HOST'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'database': os.getenv('DB_NAME')
}

# Configure Flask-Mail
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER')
app.config['MAIL_PORT'] = os.getenv('MAIL_PORT') # Convert to integer
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS') == 'True'  # Convert string to boolean
app.config['MAIL_USE_SSL'] = os.getenv('MAIL_USE_SSL') == 'True'  # Convert string to boolean
mail = Mail(app)

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('SQLALCHEMY_DATABASE_URI')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = os.getenv('SQLALCHEMY_TRACK_MODIFICATIONS') == 'False'  # Convert string to boolean
db = SQLAlchemy(app)

# Function to get a MySQL database connection
def get_db_connection():
    conn = mysql.connector.connect(**db_config)
    return conn

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

def format_date_ddmmyyyy(date_str):
    if date_str and isinstance(date_str, str) and '-' in date_str:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            return date_str  # fallback if already in correct format or invalid
    return date_str or ""

# def previous_month_range():
#     today = date.today()
#     first_day_this_month = today.replace(day=1)
#     last_day_prev_month = first_day_this_month - timedelta(days=1)
#     first_day_prev_month = last_day_prev_month.replace(day=1)
#     return first_day_prev_month, last_day_prev_month

def previous_month_range():
    return date(2025, 11, 1), date(2025, 11, 30)  # TEST MODE

def get_monthly_summary():
    start_date, end_date = previous_month_range()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            COUNT(*) AS total_invoices,
            SUM(CASE WHEN invoice_cleared = 'Yes' THEN total_amount ELSE 0 END) AS cleared_amount,
            SUM(CASE WHEN invoice_cleared = 'Yes' THEN 1 ELSE 0 END) AS cleared_count,
            SUM(CASE WHEN invoice_cleared = 'No' THEN 1 ELSE 0 END) AS uncleared_count
        FROM invoices
        WHERE invoice_date BETWEEN %s AND %s
    """, (start_date, end_date))

    summary = cursor.fetchone()

    cursor.execute("""
        SELECT invoice_date, vendor, invoice_number, total_amount, invoice_cleared
        FROM invoices
        WHERE invoice_date BETWEEN %s AND %s
        ORDER BY invoice_date
    """, (start_date, end_date))

    invoices = cursor.fetchall()

    conn.close()
    return summary, invoices, start_date, end_date


# def create_monthly_excel(summary, invoices, start_date, end_date):
#     output = BytesIO()

#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         # Sheet 1: Summary
#         summary_df = pd.DataFrame([{
#             "Month": start_date.strftime("%B %Y"),
#             "Total Invoices": summary['total_invoices'],
#             "Total Cleared Amount": summary['cleared_amount'] or 0,
#             "Cleared Invoices": summary['cleared_count'],
#             "Uncleared Invoices": summary['uncleared_count']
#         }])

#         summary_df.to_excel(writer, index=False, sheet_name="Summary")

#         # Sheet 2: Detailed Invoices
#         invoices_df = pd.DataFrame(invoices)
#         invoices_df.to_excel(writer, index=False, sheet_name="Invoices")

#     output.seek(0)
#     return output

def create_monthly_excel(summary, invoices, start_date, end_date):
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ------------------ SUMMARY SHEET ------------------
        summary_df = pd.DataFrame([{
            "Month": start_date.strftime("%B %Y"),
            "Total Invoices": summary['total_invoices'],
            "Total Cleared Amount": summary['cleared_amount'] or 0,
            "Cleared Invoices": summary['cleared_count'],
            "Uncleared Invoices": summary['uncleared_count']
        }])

        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        # ------------------ INVOICES SHEET ------------------
        invoices_df = pd.DataFrame(invoices)

        # Rename columns for business-friendly Excel
        invoices_df.rename(columns={
            'vendor': 'Vendor Name',
            'invoice_number': 'Invoice Number',
            'invoice_date': 'Invoice Date',
            'total_amount': 'Amount',
            'invoice_cleared': 'Status'
        }, inplace=True)

        # Reorder columns exactly as needed
        invoices_df = invoices_df[
            ['Vendor Name', 'Invoice Number', 'Invoice Date', 'Amount', 'Status']
        ]

        invoices_df.to_excel(writer, index=False, sheet_name="Invoices")

        # ------------------ FORMATTING ------------------
        ws = writer.sheets['Invoices']
        ws.column_dimensions['A'].width = 28  # Vendor
        ws.column_dimensions['B'].width = 20  # Invoice No
        ws.column_dimensions['C'].width = 15  # Date
        ws.column_dimensions['D'].width = 15  # Amount
        ws.column_dimensions['E'].width = 18  # Status

    output.seek(0)
    return output

def send_monthly_email(excel_file, start_date):
    subject = f"Monthly Invoice Report  {start_date.strftime('%B %Y')}"

    msg = Message(
        subject=subject,
        sender=os.getenv('MAIL_USERNAME'),
        recipients=["mihirtendulkar123@gmail.com"]  #  change
    )

    msg.body = f"""
Hello Team,

Please find attached the Monthly Invoice Report.

Month: {start_date.strftime('%B %Y')}

Regards,
Invoice Automation
    """

    msg.attach(
        filename=f"Monthly_Invoice_Report_{start_date.strftime('%b_%Y')}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=excel_file.read()
    )

    mail.send(msg)

# User model
class Users(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)  # Added user_name field
    otp = db.Column(db.String(6), nullable=True)
    role = db.Column(db.String(50), nullable=False, default='user')
    department = db.Column(db.String(100), nullable=False, default='marketing')  #  New column 
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    updated_at = db.Column(db.DateTime, default=db.func.current_timestamp(), onupdate=db.func.current_timestamp())

class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_email = db.Column(db.String(255), nullable=False)
    action = db.Column(db.String(255), nullable=False)
    department = db.Column(db.String(100), nullable=False, default='marketing')  #  New column 
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

def get_logged_in_user():
    email = request.cookies.get('email')
    if not email:
        return None
    return Users.query.filter_by(email=email.strip()).first()

# A decorator to check if the user is logged in
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not request.cookies.get('logged_in') == 'true':
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# A decorator to allow only superadmins
# Decorator to restrict access to Super Admins only
def superadmin_required(f):
    @wraps(f)
    @login_required  # This reuses your existing login check
    def decorated_function(*args, **kwargs):
        if request.cookies.get('role') != 'superadmin':
            return "Unauthorized: Access restricted to Super Admins.", 403
        return f(*args, **kwargs)
    return decorated_function

# Function to log user activity
def log_activity(action):
    user_email = request.cookies.get('email')
    if user_email:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO activity_log (user_email, action) VALUES (%s, %s)",
            (user_email, action)
        )
        conn.commit()
        conn.close()

# Endpoint route to send OTP to valid users
@app.route('/send-otp', methods=['POST'])
def send_otp():
    data = request.json
    email = data.get('email')

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    # Check if the email exists in the database
    user = Users.query.filter_by(email=email).first()
    if not user:
        return jsonify({'error': 'Email not found'}), 404

    otp = generate_otp()

    # Update the OTP in the database
    user.otp = otp
    db.session.commit()

    # Send the OTP email
    try:
        msg = Message('Your OTP Code', sender=os.getenv('MAIL_USERNAME'), recipients=[email])
        msg.body = f'Your OTP code is {otp}.'
        mail.send(msg)
    except Exception as e:
        return jsonify({'error': 'Failed to send OTP. Please try again later.'}), 500

    # Store email in cookies with expiration
    resp = make_response(jsonify({'message': 'OTP sent successfully'}))
    resp.set_cookie('email', email, max_age=60 * 60 * 4)  # Cookie expires in 4 Hours
    resp.set_cookie('role', user.role, max_age=60 * 60 * 4)  # Cookie expires in 4 Hours
    return resp

# Endpoint to check weather a valid otp is entered or not
@app.route('/verify-otp', methods=['POST'])
def verify_otp():
    data = request.json
    email = data.get('email')
    otp = data.get('otp')

    if not email or not otp:
        return jsonify({'error': 'Email and OTP are required'}), 400

    # Check if the email exists
    user = Users.query.filter_by(email=email).first()
    if not user:
        return jsonify({'valid': False}), 400

    # Ensure both OTPs are stripped of extra spaces
    user_otp = user.otp.strip() if user.otp else ''
    otp = otp.strip()

    if user_otp != otp:
        return jsonify({'valid': False}), 400

    # Clear the OTP after successful verification
    user.otp = None
    db.session.commit()

    # Set logged_in cookie
    resp = make_response(jsonify({'valid': True}))
    resp.set_cookie('logged_in', 'true', max_age=60 * 60 * 4)  # Cookie expires in 4 hours

    # Set user_name cookie
    resp.set_cookie('name', user.name, max_age=60 * 60 * 4)  # Cookie expires in 4 hours

    resp.set_cookie('role', user.role, max_age=60 * 60 * 4)  # Cookie expires in 4 hours

    return resp

# Login page route
@app.route('/')
def login():
    return render_template('login.html')

# Route to check OTP
@app.route('/otp')
def otp_page():
    # Check if the user is already logged in
    if request.cookies.get('logged_in') == 'true':
        return redirect(url_for('admin_dashboard'))
    return render_template('otp.html')

# Route for the index page ("Displays invoices")
@app.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    #  Get the logged-in user
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    query = "SELECT * FROM invoices"
    conditions = []
    params = []

    filter_type = request.args.get('filter')

    if request.method == 'POST':
        vendor = request.form.get('vendor', '').lower()
        invoice_date = request.form.get('invoice_date')
        invoice_start_date = request.form.get('invoice_start_date')
        invoice_end_date = request.form.get('invoice_end_date')
        date_submission = request.form.get('date_submission')
        invoice_number = request.form.get('invoice_number')
        po_number = request.form.get('po_number')
        created_by = request.form.get('created_by', '').lower()

        if vendor:
            conditions.append("LOWER(vendor) LIKE %s")
            params.append(f"%{vendor}%")

        if invoice_date:
            conditions.append("invoice_date = %s")
            params.append(invoice_date)
        elif invoice_start_date and invoice_end_date:
            conditions.append("invoice_date BETWEEN %s AND %s")
            params.extend([invoice_start_date, invoice_end_date])
        elif invoice_start_date:
            conditions.append("invoice_date >= %s")
            params.append(invoice_start_date)
        elif invoice_end_date:
            conditions.append("invoice_date <= %s")
            params.append(invoice_end_date)

        if invoice_number:
            conditions.append("invoice_number LIKE %s")
            params.append(f"%{invoice_number}%")

        if po_number:
            conditions.append("po_number = %s")
            params.append(po_number)

        if created_by:
            conditions.append("LOWER(created_by) LIKE %s")
            params.append(f"%{created_by}%")

    #  Add department restriction for non-superadmin users
    if user_role != 'superadmin':
        conditions.append("department = %s")
        params.append(user_department)

    #  Filter by invoice cleared status
    if filter_type == 'cleared':
        conditions.append("invoice_cleared = 'Yes'")
    elif filter_type == 'uncleared':
        conditions.append("invoice_cleared = 'No'")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY invoice_date DESC"

    cursor.execute(query, tuple(params))
    invoices = cursor.fetchall()

    no_results_message = "No matching records found." if not invoices else None

    # Dropdowns: fetch distinct values but also filtered by department for non-superadmin
    if user_role == 'superadmin':
        cursor.execute("SELECT DISTINCT vendor FROM invoices")
        vendor_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT created_by FROM invoices")
        created_by_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT invoice_date FROM invoices")
        invoice_date_values = cursor.fetchall()
    else:
        cursor.execute("SELECT DISTINCT vendor FROM invoices WHERE department = %s", (user_department,))
        vendor_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT created_by FROM invoices WHERE department = %s", (user_department,))
        created_by_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT invoice_date FROM invoices WHERE department = %s", (user_department,))
        invoice_date_values = cursor.fetchall()

    conn.close()

    return render_template(
        'index.html',
        invoices=invoices,
        filters=request.form,
        no_results_message=no_results_message,
        vendor_values=vendor_values,
        created_by_values=created_by_values,
        invoice_date_values=invoice_date_values
    )
    pass

@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
def admin_dashboard():
    filters = {}
    query = "SELECT * FROM invoices"
    params = []
    conditions = []

    #  Use helper to get user info
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    filter_type = request.args.get('filter')
    selected_month_name = request.args.get('month')
    selected_year = request.args.get('year')

    months = list(month_name)[1:]
    current_year = date.today().year
    years = list(range(current_year - 5, current_year + 1))

    if selected_month_name not in months:
        selected_month_name = month_name[date.today().month]

    try:
        selected_year = int(selected_year)
        if selected_year not in years:
            selected_year = current_year
    except (TypeError, ValueError):
        selected_year = current_year

    if request.method == 'POST':
        vendor = request.form.get('vendor', '').lower()
        invoice_date = request.form.get('invoice_date')
        invoice_start_date = request.form.get('invoice_start_date')
        invoice_end_date = request.form.get('invoice_end_date')
        invoice_number = request.form.get('invoice_number')
        po_number = request.form.get('po_number')
        created_by = request.form.get('created_by', '').lower()
        hod_approval = request.form.get('hod_approval')
        ceo_approval = request.form.get('ceo_approval')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')

        if vendor:
            conditions.append("LOWER(vendor) LIKE %s")
            params.append(f"%{vendor}%")
        if invoice_date:
            conditions.append("invoice_date = %s")
            params.append(invoice_date)
        elif invoice_start_date and invoice_end_date:
            conditions.append("invoice_date BETWEEN %s AND %s")
            params.extend([invoice_start_date, invoice_end_date])
        elif invoice_start_date:
            conditions.append("invoice_date >= %s")
            params.append(invoice_start_date)
        elif invoice_end_date:
            conditions.append("invoice_date <= %s")
            params.append(invoice_end_date)
        if invoice_number:
            conditions.append("invoice_number = %s")
            params.append(invoice_number)
        if po_number:
            conditions.append("po_number = %s")
            params.append(po_number)
        if created_by:
            conditions.append("LOWER(created_by) LIKE %s")
            params.append(f"%{created_by}%")
        if hod_approval:
            conditions.append("hod_values = %s")
            params.append(hod_approval)
        if ceo_approval:
            conditions.append("ceo_values = %s")
            params.append(ceo_approval)
        if reviewed_by:
            conditions.append("reviewed_by = %s")
            params.append(reviewed_by)

    #  Restrict department-wise data if not superadmin
    if user_role != 'superadmin':
        conditions.append("department = %s")
        params.append(user_department)

    if filter_type == 'cleared':
        conditions.append("invoice_cleared = 'Yes'")
    elif filter_type == 'uncleared':
        conditions.append("invoice_cleared = 'No'")
    elif filter_type == 'all':
        conditions = []

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY invoice_date DESC"

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(query, tuple(params))
    invoices = cursor.fetchall()
    no_results_message = "No matching records found." if not invoices else None

    # Totals & Monthly Spend Calculations
    count_query = "SELECT COUNT(*) FROM invoices"
    sum_query = 'SELECT SUM(total_amount) as total FROM invoices WHERE invoice_cleared = "Yes"'
    monthly_query = '''
        SELECT SUM(total_amount) as monthly_pool FROM invoices
        WHERE MONTH(invoice_date) = %s AND YEAR(invoice_date) = %s AND invoice_cleared = "Yes"
    '''
    past_month_query = '''
        SELECT SUM(total_amount) as past_month_spendings FROM invoices
        WHERE MONTH(invoice_date) = %s AND YEAR(invoice_date) = %s AND invoice_cleared = "Yes"
    '''

    if user_role != 'superadmin':
        count_query += " WHERE department = %s"
        sum_query += " AND department = %s"
        monthly_query += " AND department = %s"
        past_month_query += " AND department = %s"

    cursor.execute(count_query, (user_department,) if user_role != 'superadmin' else ())
    total_invoices = cursor.fetchone()['COUNT(*)']

    cursor.execute(sum_query, (user_department,) if user_role != 'superadmin' else ())
    overall_pool = cursor.fetchone()['total'] or 0

    today = date.today()
    monthly_params = (today.month, today.year) + ((user_department,) if user_role != 'superadmin' else ())
    cursor.execute(monthly_query, monthly_params)
    monthly_pool = cursor.fetchone()['monthly_pool'] or 0

    selected_month_number = months.index(selected_month_name) + 1
    past_params = (selected_month_number, selected_year) + ((user_department,) if user_role != 'superadmin' else ())
    cursor.execute(past_month_query, past_params)
    past_month_spendings = cursor.fetchone()['past_month_spendings'] or 0

    conn.close()

    return render_template(
        'admin_dashboard.html',
        invoices=invoices,
        total_invoices=total_invoices,
        total_cleared_invoices=sum(1 for i in invoices if i['invoice_cleared'] == 'Yes'),
        total_uncleared_invoices=sum(1 for i in invoices if i['invoice_cleared'] == 'No'),
        overall_pool=overall_pool,
        monthly_pool=monthly_pool,
        today=today,
        no_results_message=no_results_message,
        months=months,
        years=years,
        past_month_name=selected_month_name,
        past_year=selected_year,
        past_month_spendings=past_month_spendings
    )


# Route for download excel button("With / Without filters")
@app.route('/download_excel', methods=['POST'])
def download_excel():
    vendor = request.form.get('vendor')
    invoice_date = request.form.get('invoice_date')
    date_submission = request.form.get('date_submission')
    invoice_number = request.form.get('invoice_number')
    po_number = request.form.get('po_number')
    created_by = request.form.get('created_by')

    # Building the raw SQL query
    query = "SELECT * FROM invoices"
    conditions = []
    filters = {}

    # Apply filters based on form input
    if vendor:
        conditions.append("vendor = %s")
        filters['vendor'] = vendor
    if invoice_date:
        conditions.append("invoice_date = %s")
        filters['invoice_date'] = invoice_date
    if date_submission:
        conditions.append("date_submission = %s")
        filters['date_submission'] = date_submission
    if invoice_number:
        conditions.append("invoice_number = %s")
        filters['invoice_number'] = invoice_number
    if po_number:
        conditions.append("po_number = %s")
        filters['po_number'] = po_number
    if created_by:
        conditions.append("created_by = %s")
        filters['created_by'] = created_by

    # Append conditions to the query
    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # Execute the SQL query
    conn = get_db_connection()  # Replace this with your actual connection function
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Convert to DataFrame for easy export to Excel
    df = pd.DataFrame([{
        'Invoice Date': invoice['invoice_date'],
        'Date Received': invoice['date_received'],
        'Vendor': invoice['vendor'],
        'Invoice Number': invoice['invoice_number'],
        'PO Number': invoice['po_number'],
        'MSME': invoice['msme'],
        'Invoice Amount': invoice['invoice_amount'],
        'GST': invoice['gst'],
        'Total Amount': invoice['total_amount'],
        'Date of Submission': invoice['date_submission'],
        'Approved By': invoice['approved_by'],
        'HOD Approval': invoice['hod_values'],
        'CEO Approval': invoice['ceo_values'],
        'Reviewed By': invoice['reviewed_by'],
        'Created By': invoice['created_by'],
        'Tag1': invoice['tag1'],
        'Tag2': invoice['tag2'],
        'Invoice Cleared': invoice['invoice_cleared'],
        'Cleared Date': invoice['invoice_cleared_date'],
    } for invoice in invoices])

    # Create a bytes buffer for the Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Seek to the beginning of the stream
    output.seek(0)

    # Send the Excel file to the user for download
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='filtered_invoices.xlsx')


    # Execute the SQL query
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Check if no invoices are returned for the applied filters
    no_results_message = None
    if not invoices:
        no_results_message = "No matching records found."

    # Total number of invoices
    cursor.execute('SELECT COUNT(*) FROM invoices')
    total_invoices = cursor.fetchone()['COUNT(*)']

    # Total cleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "Yes"')
    total_cleared_invoices = cursor.fetchone()['COUNT(*)']

    # Total uncleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "No"')
    total_uncleared_invoices = cursor.fetchone()['COUNT(*)']

    conn.close()

    return render_template(
        'admin_dashboard.html', invoices=invoices,
        total_invoices=total_invoices,
        total_cleared_invoices=total_cleared_invoices,
        total_uncleared_invoices=total_uncleared_invoices,
        no_results_message=no_results_message  # Pass message to template
    )
    pass

@app.route('/add', methods=('GET', 'POST'))
@login_required
def add_invoice():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    #  Get logged-in user info
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    #  Fetch department-specific vendors
    cursor.execute('SELECT * FROM vendors WHERE department = %s', (user_department,))
    vendors = cursor.fetchall()

    #  Department-specific dropdown helper
    def get_dropdown_values(value_type):
        cursor.execute(
            "SELECT value FROM dropdown_values WHERE type = %s AND department = %s AND is_active = TRUE",
            (value_type, user_department)
        )
        return [row['value'] for row in cursor.fetchall()]
        
    # Fetch all dropdown values
    approved_by_values = get_dropdown_values('approved_by')
    created_by_values = get_dropdown_values('created_by')
    hod_values = get_dropdown_values('hod')
    ceo_values = get_dropdown_values('ceo')
    reviewed_by_values = get_dropdown_values('reviewed_by')
    tag1 = get_dropdown_values('tag1')
    tag2 = get_dropdown_values('tag2')

    if request.method == 'POST':
        # Fetch form data
        invoice_date = request.form['invoice_date']
        date_received = request.form['date_received']
        vendor = request.form['vendor']
        mobile_no = request.form['mobile_no']
        invoice_number = request.form['invoice_number']
        date_submission = request.form['date_submission']
        approved_by = request.form.get('approved_by')
        created_by = request.form['created_by']
        po_approved = request.form['po_approved']
        po_number = request.form['po_number']
        agreement_signed = request.form['agreement_signed']
        po_expiry_date = request.form.get('po_expiry_date') or None
        agreement_signed_date = request.form.get('agreement_signed_date') or None
        hod_approval = request.form.get('hod_values')
        ceo_approval = request.form.get('ceo_values')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')
        invoice_amount = float(request.form['invoice_amount'])
        gst = invoice_amount * 0.18
        total_amount = invoice_amount + gst
        total_amount = round(total_amount, 2)
        isd = request.form.get('isd', 'No')  # Default to 'No' if not selected
        invoice_cleared = request.form.get('invoice_cleared', 'No')  # Default to 'No'
        invoice_cleared_date = request.form.get('invoice_cleared_date') or None
        
        msme = request.form['msme']

        total_amount_words = num2words(total_amount, to='currency', currency='INR', lang='en_IN').upper().replace(",", "")

        # Validate form inputs
        if not invoice_date or not date_received or not vendor or not invoice_number or not date_submission or not created_by or not invoice_amount:
            flash('All fields are required!')
        else:
            # Check if the invoice number already exists
            cursor.execute('SELECT * FROM invoices WHERE invoice_number = %s', (invoice_number,))
            existing_invoice = cursor.fetchone()

            if existing_invoice:
                flash('Invoice number already exists. Please enter a unique invoice number.')
            else:
                cursor.execute(
                    '''INSERT INTO invoices (
                        invoice_date, date_received, vendor, mobile_no, invoice_number, po_approved, po_number, po_expiry_date,
                        agreement_signed, agreement_signed_date, date_submission, approved_by, created_by, tag1, tag2,
                        invoice_amount, gst, total_amount, isd, msme, hod_values, ceo_values, reviewed_by,
                        invoice_cleared, invoice_cleared_date,department
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                    (
                        invoice_date, date_received, vendor, mobile_no, invoice_number, po_approved, po_number, po_expiry_date,
                        agreement_signed, agreement_signed_date, date_submission, approved_by, created_by, tag1, tag2,
                        invoice_amount, gst, total_amount, isd, msme, hod_approval, ceo_approval, reviewed_by,
                        invoice_cleared, invoice_cleared_date,user_department
                    )
                )
                conn.commit()
                flash('Invoice added successfully!')

                user_name = request.cookies.get('name') or request.cookies.get('email')
                log_activity(f"{user_name} added invoice ({invoice_number}) for vendor {vendor}")

                # Generate Excel file using openpyxl
                #template_path = "static/excel_templates/template.xlsx"
                template_path = "static/excel_templates/Vendor_Payment_form.xlsx"

                wb = load_workbook(template_path)
                ws = wb.active

                invoice_date_fmt = format_date_ddmmyyyy(invoice_date)

                # Populate data into the template
                ws['A7'] = vendor
                ws['E6'] = invoice_date_fmt
                ws['E7'] = invoice_number
                ws['B8'] = po_approved
                ws['B9'] = agreement_signed
                ws['E8'] = po_expiry_date
                ws['E9'] = agreement_signed_date
                ws['B13'] = isd  # Or any cell where ISD info should appear
                ws['E11'] = msme
                ws['E12'] = mobile_no
                ws['F25'] = total_amount
                ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F25'].number_format = numbers.FORMAT_NUMBER_00

                ws['F35'] = total_amount
                ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F35'].number_format = numbers.FORMAT_NUMBER_00

                # ws['B35'] = total_amount_words
                if ws['A35'].value:
                    ws['A35'] = f"{ws['A35'].value} {total_amount_words}"
                else:
                    ws['A35'] = total_amount_words
                ws['A35'].alignment = Alignment(wrap_text=True)
                ws.column_dimensions['A'].width = 50 # Increase the width of column A
                ws.row_dimensions[35].height = 40  # Increase the height of row 35

                if ws['A25'].value:  # Check if there is already a value in the cell
                    ws['A25'] = f"{ws['A25'].value} ({tag1})"
                else:
                    ws['A25'] = f"Marketing Expenses ({tag1})"

                # ws['D11'] = date_submission

                if ws['D47'].value:  # Check if there's already a value in the cell
                    ws['D47'] = f"{ws['D47'].value} {created_by}"
                else:
                    ws['D47'] = created_by

                # For approved_by
                if ws['D41'].value:
                    ws['D41'] = f"{ws['D41'].value} {approved_by}"
                else:
                    ws['D41'] = approved_by

                # For reviewed_by
                if ws['D39'].value:
                    ws['D39'] = f"{ws['D39'].value} {reviewed_by}"
                else:
                    ws['D39'] = reviewed_by

                # For hod_approval
                if ws['D43'].value:
                    ws['D43'] = f"{ws['D43'].value} {hod_approval}"
                else:
                    ws['D43'] = hod_approval

                # For ceo_approval
                if ws['D45'].value:
                    ws['D45'] = f"{ws['D45'].value} {ceo_approval}"
                else:
                    ws['D45'] = ceo_approval

                # Save the populated Excel file to a BytesIO stream
                excel_file = BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                # Return the Excel file as a response
                return send_file(
                    excel_file,
                    as_attachment=True,
                    download_name=f"invoice_{invoice_number}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    conn.close()

    return render_template(
        'add_invoice.html',
        vendors=vendors,
        approved_by_values=approved_by_values,
        created_by_values=created_by_values,
        hod_values=hod_values,
        ceo_values=ceo_values,
        reviewed_by_values=reviewed_by_values,
        tag1=tag1,
        tag2=tag2
    )
    pass

@app.route('/edit/<int:id>', methods=('GET', 'POST'))
@login_required
def edit_invoice(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get the logged-in user
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    # Fetch the invoice to be edited
    cursor.execute('SELECT * FROM invoices WHERE id = %s', (id,))
    invoice = cursor.fetchone()

    if not invoice:
        flash("Invoice not found.")
        return redirect(url_for('dashboard'))

    # Fetch vendors based on department
    cursor.execute('SELECT * FROM vendors WHERE department = %s', (user_department,))
    vendors = cursor.fetchall()

    # Helper to fetch dropdown values department-wise
    def get_dropdown_values(value_type):
        cursor.execute(
            "SELECT value FROM dropdown_values WHERE type = %s AND department = %s AND is_active = TRUE",
            (value_type, user_department)
        )
        return [row['value'] for row in cursor.fetchall()]

    approved_by_values = get_dropdown_values('approved_by')
    created_by_values = get_dropdown_values('created_by')
    hod_values = get_dropdown_values('hod')
    ceo_values = get_dropdown_values('ceo')
    reviewed_by_values = get_dropdown_values('reviewed_by')
    tag1 = get_dropdown_values('tag1')
    tag2 = get_dropdown_values('tag2')

    if request.method == 'POST':
        invoice_date = request.form['invoice_date']
        date_received = request.form['date_received']
        vendor = request.form['vendor']
        mobile_no = request.form['mobile_no']
        invoice_number = request.form['invoice_number']
        po_approved = request.form['po_approved']
        po_number = request.form['po_number']
        agreement_signed = request.form['agreement_signed']
        date_submission = request.form['date_submission']
        isd = request.form.get('isd')  # Default to 'No' if not selected
        approved_by = request.form.get('approved_by')
        hod_approval = request.form.get('hod_values')
        ceo_approval = request.form.get('ceo_values')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')
        po_expiry_date = request.form.get('po_expiry_date')  # Optional
        agreement_signed_date = request.form.get('agreement_signed_date')  # Optional
        po_expiry_date = po_expiry_date if po_expiry_date else None
        agreement_signed_date = agreement_signed_date if agreement_signed_date else None
        created_by = invoice['created_by']
        invoice_cleared = request.form['invoice_cleared']

        if invoice_cleared == 'Yes':
            invoice_cleared_date = request.form.get('invoice_cleared_date') or date.today()
        else:
            invoice_cleared_date = None

        invoice_amount = float(request.form['invoice_amount'])
        gst = invoice_amount * 0.18
        total_amount = invoice_amount + gst
        total_amount_words = num2words(total_amount, to='currency', currency='INR', lang='en_IN')
        total_amount_words = total_amount_words.upper().replace(",", "")
        msme = request.form['msme']

        if not invoice_date or not date_received or not vendor or not invoice_number or not date_submission or not invoice_amount:
            flash('All fields are required!')
        else:
            cursor.execute(
                '''
                UPDATE invoices 
                SET invoice_date = %s, 
                    date_received = %s, 
                    vendor = %s, 
                    mobile_no = %s,
                    invoice_number = %s, 
                    po_approved = %s, 
                    po_number = %s, 
                    po_expiry_date = %s, 
                    agreement_signed = %s, 
                    agreement_signed_date = %s, 
                    date_submission = %s, 
                    approved_by = %s, 
                    created_by = %s, 
                    tag1 = %s, 
                    tag2 = %s, 
                    invoice_amount = %s, 
                    gst = %s,
                    isd = %s, 
                    total_amount = %s, 
                    msme = %s, 
                    hod_values = %s, 
                    ceo_values = %s, 
                    reviewed_by = %s, 
                    invoice_cleared = %s, 
                    invoice_cleared_date = %s
                WHERE id = %s
                ''',
                (
                    invoice_date, date_received, vendor, mobile_no, invoice_number, 
                    po_approved, po_number, po_expiry_date, agreement_signed, 
                    agreement_signed_date, date_submission, approved_by, created_by, 
                    tag1, tag2, invoice_amount, gst,isd ,total_amount, msme, 
                    hod_approval, ceo_approval, reviewed_by, invoice_cleared, 
                    invoice_cleared_date, id
                )
            )
            conn.commit()

            flash('Invoice updated successfully!')

            user_name = request.cookies.get('name') or request.cookies.get('email')
            log_activity(f"{user_name} edited invoice ({invoice_number}) for vendor {vendor}")

            # Skip Excel generation if invoice is cleared
            if invoice_cleared == 'Yes':
                conn.close()
                return redirect(url_for('index'))  # Redirect to a different view after successful update
            
            # Generate the updated invoice Excel
            # template_path = "static/excel_templates/template.xlsx"
            template_path = "static/excel_templates/Vendor_Payment_form.xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            invoice_date_fmt = format_date_ddmmyyyy(invoice_date)

            # Populate data into the template
            ws['A7'] = vendor
            ws['E6'] = invoice_date_fmt
            ws['E7'] = invoice_number
            ws['B8'] = po_approved
            ws['B9'] = agreement_signed
            ws['E8'] = po_expiry_date
            ws['E9'] = agreement_signed_date
            ws['E11'] = msme
            ws['B13'] = isd  # Or any cell where ISD info should appear
            ws['E12'] = mobile_no
            ws['F25'] = total_amount
            ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F25'].number_format = numbers.FORMAT_NUMBER_00

            ws['F35'] = total_amount
            ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F35'].number_format = numbers.FORMAT_NUMBER_00

            if ws['A35'].value:
                ws['A35'] = f"{ws['A35'].value} {total_amount_words}"
            else:
                ws['A35'] = total_amount_words
            ws['A35'].alignment = Alignment(wrap_text=True)
            ws.column_dimensions['A'].width = 50  # Increase the width of column A
            ws.row_dimensions[35].height = 40  # Increase the height of row 35

            if ws['A25'].value:  # Check if there is already a value in the cell
                ws['A25'] = f"{ws['A25'].value} ({tag1})"
            else:
                ws['A25'] = f"Marketing Expenses ({tag1})"

            if ws['D47'].value:  # Check if there's already a value in the cell
                ws['D47'] = f"{ws['D47'].value} {created_by}"
            else:
                ws['D47'] = created_by

            # For approved_by
            if ws['D41'].value:
                ws['D41'] = f"{ws['D41'].value} {approved_by}"
            else:
                ws['D41'] = approved_by

            # For reviewed_by
            if ws['D39'].value:
                ws['D39'] = f"{ws['D39'].value} {reviewed_by}"
            else:
                ws['D39'] = reviewed_by

            # For hod_approval
            if ws['D43'].value:
                ws['D43'] = f"{ws['D43'].value} {hod_approval}"
            else:
                ws['D43'] = hod_approval

            # For ceo_approval
            if ws['D45'].value:
                ws['D45'] = f"{ws['D45'].value} {ceo_approval}"
            else:
                ws['D45'] = ceo_approval
            
            # Save the populated Excel file to a BytesIO stream
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Return the Excel file as a response
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=f"updated_invoice_{invoice_number}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    conn.close()

    return render_template(
        'edit_invoice.html', 
        invoice=invoice, 
        vendors=vendors, 
        approved_by_values=approved_by_values, 
        created_by_values=created_by_values,
        hod_values=hod_values,
        ceo_values=ceo_values,
        reviewed_by_values=reviewed_by_values,
        tag1=tag1,
        tag2=tag2
    )
    pass

@app.route('/delete/<int:id>', methods=('POST',))
@login_required
def delete_invoice(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get the logged-in user
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in.")
        return redirect(url_for('login'))

    user_name = user.name or user.email  # Fallback
    user_department = user.department
    user_role = user.role

    # Fetch the invoice before deletion
    cursor.execute('SELECT invoice_number, vendor, department FROM invoices WHERE id = %s', (id,))
    invoice = cursor.fetchone()

    # Optional: Restrict deletion by department or role
    if invoice:
        if user_role != 'admin' and invoice['department'] != user_department:
            flash("You are not authorized to delete this invoice.")
            return redirect(url_for('index'))

        # Delete the invoice
        cursor.execute('DELETE FROM invoices WHERE id = %s', (id,))
        conn.commit()
        log_activity(f"{user_name} deleted invoice ({invoice['invoice_number']}) for vendor {invoice['vendor']}")
        flash('Invoice deleted successfully!')
    else:
        flash("Invoice not found.")

    conn.close()
    return redirect(url_for('index'))

@app.route('/manage_vendors', methods=['GET', 'POST'])
@superadmin_required
def manage_vendors():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        new_vendor = request.form['new_vendor']
        vendor_status = request.form['vendor_status']  # e.g., 'Active' or 'Inactive'

        if new_vendor:
            cursor.execute(
                'INSERT INTO vendors (vendor_name, vendor_status) VALUES (%s, %s)',
                (new_vendor, vendor_status)
            )
            conn.commit()
            flash('Vendor added successfully!')

            #  Log the activity
            user_name = request.cookies.get('name') or request.cookies.get('email')
            log_activity(f"{user_name} added vendor '{new_vendor}' with status '{vendor_status}'")

        else:
            flash('Vendor name cannot be empty!')

    # Fetch all vendors
    cursor.execute('SELECT * FROM vendors')
    vendors = cursor.fetchall()
    conn.close()

    return render_template('manage_vendors.html', vendors=vendors)


@app.route('/edit_vendor/<int:id>', methods=['POST'])
@superadmin_required
def edit_vendor(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    vendor_status = request.form['vendor_status']

    # Fetch vendor name before update (for logging)
    cursor.execute('SELECT vendor_name FROM vendors WHERE id = %s', (id,))
    vendor = cursor.fetchone()
    vendor_name = vendor['vendor_name'] if vendor else 'Unknown Vendor'

    # Perform the update
    cursor.execute('UPDATE vendors SET vendor_status = %s WHERE id = %s', (vendor_status, id))
    conn.commit()
    conn.close()

    #  Log the activity
    user_name = request.cookies.get('name') or request.cookies.get('email')
    log_activity(f"{user_name} updated vendor '{vendor_name}' status to '{vendor_status}'")

    flash('Vendor status updated successfully!')
    return redirect(url_for('manage_vendors'))


# Route to manage dropdowns (add new values)
@app.route('/manage_dropdowns', methods=['GET', 'POST'])
@superadmin_required
def manage_dropdowns():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        type_ = request.form.get('type')
        value = request.form.get('value')
        if type_ and value:
            cursor.execute(
                'INSERT INTO dropdown_values (type, value) VALUES (%s, %s)', 
                (type_, value)
            )
            conn.commit()
            flash('Value added successfully.')

            #  Log the activity
            user_name = request.cookies.get('name') or request.cookies.get('email')
            log_activity(f"{user_name} added a new value '{value}' to the '{type_}' dropdown")

    # Load all dropdowns
    cursor.execute('SELECT DISTINCT type FROM dropdown_values')
    types = [row['type'] for row in cursor.fetchall()]

    dropdown_data = {}
    for t in types:
        cursor.execute(
            'SELECT id, value FROM dropdown_values WHERE type = %s AND is_active = TRUE', 
            (t,)
        )
        dropdown_data[t] = cursor.fetchall()

    conn.close()
    return render_template('manage_dropdowns.html', dropdown_data=dropdown_data)


#Delete dropdown value from Database(soft delete)
@app.route('/delete_dropdown/<int:id>', methods=['POST'])
@superadmin_required
def delete_dropdown(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch the value before updating for logging
    cursor.execute('SELECT type, value FROM dropdown_values WHERE id = %s', (id,))
    value_data = cursor.fetchone()

    if value_data:
        cursor.execute('UPDATE dropdown_values SET is_active = FALSE WHERE id = %s', (id,))
        conn.commit()

        #  Log the activity
        user_name = request.cookies.get('name') or request.cookies.get('email')
        log_activity(f"{user_name} soft-deleted the value '{value_data['value']}' from '{value_data['type']}' dropdown")

        flash('Value disabled (soft-deleted).')
    else:
        flash('Dropdown value not found.')

    conn.close()
    return redirect(url_for('manage_dropdowns'))


# Route to manage users (superadmin only)
@app.route('/manage_users')
@superadmin_required
def manage_users():
    users = Users.query.all()

    #  Log the activity
    user_name = request.cookies.get('name') or request.cookies.get('email')
    log_activity(f"{user_name} viewed the Manage Users page")

    return render_template('manage_users.html', users=users)

# Route to add a new user (superadmin only)
@app.route('/add_user', methods=['POST'])
@superadmin_required
def add_user():
    name = request.form['name']
    email = request.form['email']
    role = request.form['role']

    existing_user = Users.query.filter_by(email=email).first()
    if not existing_user:
        new_user = Users(name=name, email=email, role=role, is_active=True)
        db.session.add(new_user)
        db.session.commit()

        #  Log activity
        actor = request.cookies.get('name') or request.cookies.get('email')
        log_activity(f"{actor} added new user: {name} ({email}) with role {role}")

    return redirect(url_for('manage_users'))

# Route to edit/delete user details (superadmin only)
@app.route('/delete_user/<int:user_id>', methods=['POST'])
@superadmin_required
def delete_user(user_id):
    user = Users.query.get_or_404(user_id)

    #  Get actor BEFORE deletion
    actor = request.cookies.get('name') or request.cookies.get('email')

    #  Log BEFORE deletion
    log_activity(f"{actor} deleted user: {user.name} ({user.email})")

    db.session.delete(user)
    db.session.commit()

    return redirect(url_for('manage_users'))


# Route to toggle user status (active/inactive) (superadmin only)
@app.route('/toggle_user_status/<int:user_id>', methods=['POST'])
@superadmin_required
def toggle_user_status(user_id):
    user = Users.query.get_or_404(user_id)

    # Toggle the status
    user.is_active = not user.is_active
    db.session.commit()

    # Identify actor from cookies
    actor = request.cookies.get('name') or request.cookies.get('email')

    # Log the action
    status = "activated" if user.is_active else "deactivated"
    log_activity(f"{actor} {status} user: {user.name} ({user.email})")

    return redirect(url_for('manage_users'))


# Route to update user role (superadmin only)
@app.route('/update_user_role/<int:user_id>', methods=['POST'])
@superadmin_required
def update_user_role(user_id):
    user = Users.query.get_or_404(user_id)
    new_role = request.form.get('role')

    if new_role in ['user', 'admin', 'superadmin']:
        old_role = user.role
        user.role = new_role
        db.session.commit()

        # Identify actor from cookie
        actor = request.cookies.get('name') or request.cookies.get('email')

        # Log the role change
        log_activity(f"{actor} changed role of user {user.name} ({user.email}) from {old_role} to {new_role}")

    return redirect(url_for('manage_users'))

# Route to view activity logs (superadmin only)
@app.route('/activity_logs')
@superadmin_required
def view_activity_logs():
    # Only show logs from the last 15 days
    fifteen_days_ago = datetime.utcnow() - timedelta(days=15)
    logs = ActivityLog.query.filter(ActivityLog.timestamp >= fifteen_days_ago).order_by(ActivityLog.timestamp.desc()).all()

    return render_template('activity_logs.html', logs=logs)

# Route to download activity logs as CSV (superadmin only)
@app.route('/download_activity_logs')
@superadmin_required
def download_activity_logs():
    # Use StringIO to write text data
    string_io = StringIO()
    writer = csv.writer(string_io)
    writer.writerow(['ID', 'User Email', 'Action', 'Timestamp'])

    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).all()
    for log in logs:
        writer.writerow([log.id, log.user_email, log.action, log.timestamp.strftime('%Y-%m-%d %H:%M:%S')])

    # Convert string data to bytes
    mem = BytesIO()
    mem.write(string_io.getvalue().encode('utf-8'))  # encode to bytes
    mem.seek(0)

    filename = f"activity_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(mem,
                     download_name=filename,
                     mimetype='text/csv',
                     as_attachment=True)

# Route for user login
@app.route('/logout')
def logout():
    # Clear session or cookies here
    flash("Logged out successfully.")
    return redirect(url_for('login'))

# Middleware to add headers to prevent caching
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

@app.route('/test_monthly_report')
@login_required
def test_monthly_report():
    run_monthly_report()
    return "Monthly report sent successfully (TEST MODE)"

def run_monthly_report():
    with app.app_context():
        summary, invoices, start_date, end_date = get_monthly_summary()
        excel = create_monthly_excel(summary, invoices, start_date, end_date)
        send_monthly_email(excel, start_date)

        # Optional log
        log_activity(f"System sent monthly report for {start_date.strftime('%B %Y')}")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
