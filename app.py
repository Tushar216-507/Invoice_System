from flask import Flask, jsonify, render_template, request, redirect, url_for, flash, make_response
import pandas as pd
from io import BytesIO
from flask import send_file 
import mysql.connector
from datetime import date, datetime, timedelta, timezone
import os
import string
from flask_sqlalchemy import SQLAlchemy
import random
from dotenv import load_dotenv
from flask_mail import Mail, Message
from functools import wraps
try:
    from xhtml2pdf import pisa
    XHTML2PDF_AVAILABLE = True
except ImportError:
    pisa = None
    XHTML2PDF_AVAILABLE = False
    print("Warning: xhtml2pdf not available - PDF generation will be disabled")
from num2words import num2words
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from calendar import month_name
import calendar
import csv
from io import StringIO, BytesIO
import json   #new_import   
from openai import OpenAI  #new_import
#Imports for PO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from reportlab.lib.units import inch
from utils import amount_to_words
#-------------------------For logo in PO------------------------------------------
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Frame
from PIL import Image
Image.MAX_IMAGE_PIXELS = None
import io
#-------------------------------New Security Update------------------------------------
from functools import wraps
from flask_login import LoginManager, login_user, logout_user, current_user, UserMixin, login_required
from flask_wtf.csrf import CSRFProtect
from flask import abort
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask import session
from datetime import timedelta
from flask_login import current_user
import secrets
#-------------------------------Chatbot required imports------------------------------------
load_dotenv()
#-------------------------------WhatsApp Integration (DICE API)------------------------------------
import base64
import requests
# from backend.new_chatbot import chatbot
# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================

import logging
from logging.handlers import RotatingFileHandler
import os

# Configure logging based on environment
if os.getenv('FLASK_ENV') == 'production':
    log_level = logging.WARNING
else:
    log_level = logging.INFO

# Create logs directory if it doesn't exist
if not os.path.exists('logs'):
    os.makedirs('logs')

# Configure file handler with rotation
file_handler = RotatingFileHandler(
    'logs/invoice_app.log',
    maxBytes=10485760,  # 10MB
    backupCount=10
)
file_handler.setLevel(log_level)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))

# Configure console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(log_level)
console_handler.setFormatter(logging.Formatter(
    '%(levelname)s: %(message)s'
))

# Get the app logger
app_logger = logging.getLogger('invoice_app')
app_logger.setLevel(log_level)
app_logger.addHandler(file_handler)
app_logger.addHandler(console_handler)

logger = app_logger

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ALLOWED_INTENTS = {
    "monthly_spend",
    "uncleared_invoices",
    "vendor_summary",
    "invoice_list",
    "general_summary",
    "cleared_invoices",
    "total_invoices",
    "invoice_details",
    "vendor_details"
}

app = Flask(__name__)
app.secret_key = os.getenv('APP_SECRET_KEY')  # Secret key for flash messages
# ---- SECURITY FRAMEWORK ----
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.refresh_view = "login"
login_manager.needs_refresh_message = "Session expired. Please login again."
login_manager.needs_refresh_message_category = "info"

csrf = CSRFProtect(app)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"]
)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"]
)

def get_tag1_monthly_trends(selected_fy, trend_tag=None):
    """
    Get monthly spending trends for Tag1 categories
    
    Args:
        selected_fy: Financial year string (e.g., "2024-2025")
        trend_tag: Optional specific tag1 to filter (None = all tags)
    
    Returns:
        Dictionary with 'labels' (months) and 'data' (spending per tag)
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Parse financial year
    start_year = int(selected_fy.split('-')[0])
    
    # Financial year runs from April to March
    fy_start = f"{start_year}-04-01"
    fy_end = f"{start_year + 1}-03-31"
    
    # Base query
    if trend_tag:
        # Single tag query
        query = """
            SELECT 
                DATE_FORMAT(invoice_date, '%Y-%m') as month,
                tag1,
                SUM(total_amount) as total
            FROM invoices
            WHERE invoice_date BETWEEN %s AND %s
                AND tag1 = %s
            GROUP BY DATE_FORMAT(invoice_date, '%Y-%m'), tag1
            ORDER BY month
        """
        cursor.execute(query, (fy_start, fy_end, trend_tag))
    else:
        # All tags query - get top 6 tags by total spending
        query = """
            SELECT 
                DATE_FORMAT(invoice_date, '%Y-%m') as month,
                tag1,
                SUM(total_amount) as total
            FROM invoices
            WHERE invoice_date BETWEEN %s AND %s
                AND tag1 IS NOT NULL
                AND tag1 != ''
                AND tag1 IN (
                    SELECT tag1 
                    FROM (
                        SELECT tag1, SUM(total_amount) as tag_total
                        FROM invoices
                        WHERE invoice_date BETWEEN %s AND %s
                            AND tag1 IS NOT NULL
                            AND tag1 != ''
                        GROUP BY tag1
                        ORDER BY tag_total DESC
                        LIMIT 6
                    ) as top_tags
                )
            GROUP BY DATE_FORMAT(invoice_date, '%Y-%m'), tag1
            ORDER BY month, tag1
        """
        cursor.execute(query, (fy_start, fy_end, fy_start, fy_end))
    
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    
    # Generate all months in the financial year
    from datetime import datetime
    
    months = []
    current = datetime(start_year, 4, 1)  # April 1st
    end = datetime(start_year + 1, 4, 1)  # Next April 1st
    
    while current < end:
        months.append(current.strftime('%Y-%m'))
        # Move to next month
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    
    # Format labels as "MMM YYYY"
    labels = [datetime.strptime(m, '%Y-%m').strftime('%b %Y') for m in months]
    
    # Organize data by tag
    tag_data = {}
    for row in results:
        tag = row['tag1']
        month = row['month']
        total = float(row['total'])
        
        if tag not in tag_data:
            tag_data[tag] = {m: 0 for m in months}
        
        if month in tag_data[tag]:
            tag_data[tag][month] = total
    
    # Convert to array format for Chart.js
    data = {}
    for tag, month_totals in tag_data.items():
        data[tag] = [month_totals[m] for m in months]
    
    return {
        'labels': labels,
        'data': data
    }

def get_vendor_monthly_trends(selected_fy, trend_vendor=None):
    """
    Get monthly spending trends for vendors
    
    Args:
        selected_fy: Financial year string (e.g., "2024-2025")
        trend_vendor: Optional specific vendor to filter (None = top 6 vendors)
    
    Returns:
        Dictionary with 'labels' (months) and 'data' (spending per vendor)
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Parse financial year
    start_year = int(selected_fy.split('-')[0])
    
    # Financial year runs from April to March
    fy_start = f"{start_year}-04-01"
    fy_end = f"{start_year + 1}-03-31"
    
    # Base query
    if trend_vendor:
        # Single vendor query
        query = """
            SELECT 
                DATE_FORMAT(invoice_date, '%Y-%m') as month,
                vendor,
                SUM(total_amount) as total
            FROM invoices
            WHERE invoice_date BETWEEN %s AND %s
                AND vendor = %s
                AND invoice_cleared = 'Yes'
            GROUP BY DATE_FORMAT(invoice_date, '%Y-%m'), vendor
            ORDER BY month
        """
        cursor.execute(query, (fy_start, fy_end, trend_vendor))
    else:
        # All vendors query - get top 6 vendors by total spending
        query = """
            SELECT 
                DATE_FORMAT(invoice_date, '%Y-%m') as month,
                vendor,
                SUM(total_amount) as total
            FROM invoices
            WHERE invoice_date BETWEEN %s AND %s
                AND invoice_cleared = 'Yes'
                AND vendor IS NOT NULL
                AND vendor != ''
                AND vendor IN (
                    SELECT vendor 
                    FROM (
                        SELECT vendor, SUM(total_amount) as vendor_total
                        FROM invoices
                        WHERE invoice_date BETWEEN %s AND %s
                            AND invoice_cleared = 'Yes'
                            AND vendor IS NOT NULL
                            AND vendor != ''
                        GROUP BY vendor
                        ORDER BY vendor_total DESC
                        LIMIT 6
                    ) as top_vendors
                )
            GROUP BY DATE_FORMAT(invoice_date, '%Y-%m'), vendor
            ORDER BY month, vendor
        """
        cursor.execute(query, (fy_start, fy_end, fy_start, fy_end))
    
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    
    # Generate all months in the financial year
    from datetime import datetime
    
    months = []
    current = datetime(start_year, 4, 1)  # April 1st
    end = datetime(start_year + 1, 4, 1)  # Next April 1st
    
    while current < end:
        months.append(current.strftime('%Y-%m'))
        # Move to next month
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)
    
    # Format labels as "MMM YYYY"
    labels = [datetime.strptime(m, '%Y-%m').strftime('%b %Y') for m in months]
    
    # Organize data by vendor
    vendor_data = {}
    for row in results:
        vendor = row['vendor']
        month = row['month']
        total = float(row['total'])
        
        if vendor not in vendor_data:
            vendor_data[vendor] = {m: 0 for m in months}
        
        if month in vendor_data[vendor]:
            vendor_data[vendor][month] = total
    
    # Convert to array format for Chart.js
    data = {}
    for vendor, month_totals in vendor_data.items():
        data[vendor] = [month_totals[m] for m in months]
    
    return {
        'labels': labels,
        'data': data
    }

@app.before_request
def refresh_session():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(hours=4)

# Secure Session Cookies
from datetime import timedelta
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Strict"
app.config["SESSION_COOKIE_SECURE"] = False  # ⚠️ Change to True in HTTPS
app.config["REMEMBER_COOKIE_DURATION"] = timedelta(days=7)
app.config["REMEMBER_COOKIE_REFRESH_EACH_REQUEST"] = True
app.config["SESSION_PROTECTION"] = "strong"

app.permanent_session_lifetime = timedelta(hours=4)

db_config = {
    'host': os.getenv('DB_HOST'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'database': os.getenv('DB_NAME')
}

from mysql.connector import pooling

try:
    db_pool = pooling.MySQLConnectionPool(
        pool_name="invoice_pool",
        pool_size=10,
        pool_reset_session=True,
        **db_config
    )
    logger.info("✅ Database connection pool created")
except Exception as e:
    logger.error(f"❌ Failed to create connection pool: {e}")
    db_pool = None

LOGO_PATH = "static/logo.png"
PAGE_WIDTH, PAGE_HEIGHT = A4

def add_page_layout(canvas, doc):
    try:
        from PIL import Image
        Image.MAX_IMAGE_PIXELS = None       # allow large PNG

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(BASE_DIR, "static", "logo.png")

        img = Image.open(logo_path).convert("RGBA")

        # ----- BIG NICE LETTERHEAD LOGO -----
        desired_width_pt = 3.0 * inch        # ← make bigger/smaller here
        ratio = img.height / img.width
        desired_height_pt = desired_width_pt * ratio

        img_byte = io.BytesIO()
        img.save(img_byte, format="PNG")
        img_byte.seek(0)

        logo = ImageReader(img)

        canvas.saveState()

        x = A4[0] - desired_width_pt - 10     # 35 px right margin
        y = A4[1] - desired_height_pt + 40   # 35 px top margin

        canvas.drawImage(
            logo,
            x, y,
            width=desired_width_pt,
            height=desired_height_pt,
            preserveAspectRatio=True,
            mask='auto'
        )

    except Exception as e:
        logger.error(f"Logo rendering error: {e}")

    # ================= FOOTER =================
    footer_html = """
<b>Regd. and Corporate Office</b><br/>
<b>Auxilo Finserve Private Limited</b><br/>
Office No. 552, 6th Floor, Kalpataru Square,<br/>
Kondivita Road, Andheri East, Mumbai 400059, Maharashtra, India<br/>
✆: +91 22 62463333 &nbsp;&nbsp; ✉: support@auxilo.com &nbsp;&nbsp; 🌐:www.auxilo.com<br/>
<b>CIN No:</b> U65990MH2016PTC282516
"""

    style = ParagraphStyle(
        name="FooterStyle",
        fontName="Times-Roman",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
        textColor=colors.black
    )

    para = Paragraph(footer_html, style)

    # X , Y , Width , Height
    frame = Frame(40, -30, 500, 120, showBoundary=0)
    frame.addFromList([para], canvas)
# Function for Pdf layout for PO
def generate_po_pdf_flask(data):
    """
    Generate Purchase Order PDF - Everything in ONE single table
    """
    os.makedirs("generated_pdfs", exist_ok=True)
    po_num = data.get('po_number', '')
    if po_num and po_num != 'N/A':
        safe_name = po_num.replace('/', '_')
    else:
        safe_name = f"PO_NoNumber_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    file_path = f"generated_pdfs/{safe_name}.pdf"

    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=2*inch,
        bottomMargin=0.5*inch
    )

    # ==================== STYLES ====================
    title_style = ParagraphStyle(
        'Title',
        fontName='Times-Bold',
        fontSize=18,
        leading=22,
        alignment=TA_CENTER
    )

    normal_style = ParagraphStyle(
        'Normal',
        fontName='Times-Roman',
        fontSize=10,
        leading=12
    )

    bold_style = ParagraphStyle(
        'Bold',
        fontName='Times-Bold',
        fontSize=10,
        leading=12
    )

    small_style = ParagraphStyle(
        'Small',
        fontName='Times-Roman',
        fontSize=9,
        leading=11
    )

    # ==================== BUILD ONE MASSIVE TABLE ====================
    all_rows = []

    # ROW 1: TITLE (PURCHASE ORDER) - Spans all columns
    all_rows.append([Paragraph("PURCHASE ORDER", title_style), "", "", "", "", "", ""])

    # ROW 2: Header (To, Billing Address, PO-NO)
    # Left: Vendor Address only (no vendor name)
    # Middle: Company Billing Address with GST
    # Right: PO Number and Date (GST removed from here)
    
    billing_address_text = (
        "<b>BILLING ADDRESS:</b><br/>"
        "Auxilo Finserve Pvt Ltd<br/>"
        "Office no 63, 06th Floor,<br/>"
        "Kalpataru Square, Kondivita Road, Andheri East,<br/>"
        "Mumbai 400059.<br/><br/>"
        f"<b>GST:</b> 27AAXCS7051B1Z2"
    )
    
    all_rows.append([
        Paragraph(f"<b>To,</b><br/>{data['vendor_address']}", normal_style),
        "",  # Merge
        Paragraph(billing_address_text, normal_style),
        "",  # Merge
        Paragraph(
            f"<b>PO-NO.</b><br/>{data.get('po_number')}<br/>"
            f"<b>Date:</b> {data['date']}", 
            normal_style
        ),
        "",  # Merge
        ""   # Merge
    ])

    # ROW 3: Items table header
    all_rows.append([
        Paragraph("<b>Sr.<br/>No</b>", bold_style),
        Paragraph("<b>Product Description</b>", bold_style),
        "",  # Merge with description
        Paragraph("<b>Qty</b>", bold_style),
        Paragraph("<b>Rate</b>", bold_style),
        Paragraph("<b>CGST<br/>9%</b>", bold_style),
        Paragraph("<b>SGST<br/>9%</b>", bold_style),
        Paragraph("<b>Total<br/>(INR)</b>", bold_style)
    ])

    # ROW 4+: Item rows
    if data["items"]:
        for i, item in enumerate(data["items"], 1):
            qty_str = str(int(item['qty'])) if item['qty'] == int(item['qty']) else f"{item['qty']:.1f}"
            
            all_rows.append([
                Paragraph(f"{i}.", normal_style),
                Paragraph(item["description"], normal_style),
                "",  # Merge
                Paragraph(qty_str, normal_style),
                Paragraph(f"{item['rate']:,.2f}", normal_style),
                Paragraph(f"{item['cgst']:,.2f}", normal_style),
                Paragraph(f"{item['sgst']:,.2f}", normal_style),
                Paragraph(f"{item['total']:,.2f}", normal_style)
            ])

    # ROW: Total
    total_row_index = len(all_rows)
    all_rows.append([
        "",
        Paragraph("<b>Total</b>", bold_style),
        "",
        "",
        "",
        "",
        "",
        Paragraph(f"<b>{data['grand_total']:,.2f}</b>", bold_style)
    ])

    # ROW: Amount in words
    words_row_index = len(all_rows)
    all_rows.append([
    Paragraph(f"<b>In Words:</b> {data.get('amount_words', '')}", normal_style),
    "", "", "", "", "", "", ""
    ])

    # ROW: Footer (Terms and Company details)
    terms = (
        "<b>*THIS COST IS FOR ONE TIME ONLY</b><br/><br/>"
        "1. Purchase order is inclusive of all taxes and<br/>"
        "delivery terms: 4 to 5 working days from receipt of Purchase<br/>"
        "Order.<br/><br/>"
        "2. Please send two copies of your invoice.<br/><br/>"
        "3. Enter this order in accordance with the prices, terms, method<br/>"
        "and listed above specified and payment terms :50% Payment<br/>"
        "advance and balance after delivery.<br/><br/>"
        "4. Please notify us immediately if you are unable to ship as<br/>"
        "specified.<br/><br/>"
        "5. We reserve the right to reject goods that are not in good order<br/>"
        "or condition as determined by our quality control.<br/><br/>"
        "6. Send all correspondence to: marketing@auxilo.com."
    )

    # Build company details as simple text (no nested table grid)
    company_text = (
        f"<b>PAN No.</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {data.get('pan_no', 'AAXCS7051B')}<br/><br/><br/>"
        f"<b>GSTI No</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: 27AAXCS7051B1Z2<br/><br/><br/>"
        f"<b>Signature</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:<br/><br/>"
        f"<b>Name</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: Benoy Joseph<br/>"
        f"<b>Designation</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: Head – Marketing<br/><br/><br/>"
        f"<b>Date</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {data['date']}<br/><br/>"
        f"<b>Place</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: Mumbai"
    )
    
    company_table = Paragraph(company_text, normal_style)

    footer_row_index = len(all_rows)
    all_rows.append([
        Paragraph(terms, small_style),
        "", "", "",
        company_table,
        "", "", ""
    ])

    # ==================== CREATE THE MAIN TABLE ====================
    main_table = Table(
        all_rows,
        colWidths=[0.4*inch, 1.4*inch, 1.4*inch, 0.5*inch, 1.0*inch, 0.8*inch, 0.8*inch, 1.0*inch]
    )

    # ==================== APPLY ALL STYLES ====================
    style_commands = [
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        
        # ROW 1: Title - Span all columns, center align
        ('SPAN', (0,0), (7,0)),
        ('ALIGN', (0,0), (0,0), 'CENTER'),
        
        # ROW 2: Header - 3 columns span
        ('SPAN', (0,1), (1,1)),  # To (Vendor Address)
        ('SPAN', (2,1), (3,1)),  # Billing Address
        ('SPAN', (4,1), (7,1)),  # PO-NO
        ('VALIGN', (0,1), (-1,1), 'TOP'),
        
        # ROW 3: Items header
        ('SPAN', (1,2), (2,2)),  # Product Description spans 2 columns
        ('ALIGN', (0,2), (-1,2), 'CENTER'),
        
        # Items data rows alignment
        ('ALIGN', (0,3), (0,-1), 'LEFT'),      # Sr No left
        ('ALIGN', (3,3), (7,-1), 'CENTER'),    # Numbers center
    ]

    # Total row span
    style_commands.append(('SPAN', (1, total_row_index), (6, total_row_index)))
    
    # Words row span
    style_commands.append(('SPAN', (0, words_row_index), (7, words_row_index)))
    
    # Footer row span
    style_commands.append(('SPAN', (0, footer_row_index), (3, footer_row_index)))  # Left column
    style_commands.append(('SPAN', (4, footer_row_index), (7, footer_row_index)))  # Right column
    style_commands.append(('VALIGN', (0, footer_row_index), (-1, footer_row_index), 'TOP'))

    # Padding
    style_commands.extend([
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ])

    main_table.setStyle(TableStyle(style_commands))

    # Build PDF with just the one table
    doc.build(
    [main_table],
    onFirstPage=add_page_layout,
    onLaterPages=add_page_layout
    )
    return file_path

#------------------------------- ----------------------------------------
# CHATBOT ENDPOINT (Legacy - redirects to new chatbot)
#----------------------------------------------------------------------------
@app.route('/api/chat', methods=['POST'])
@login_required
def chat_endpoint():
    """
    Legacy chatbot endpoint - uses new chatbot from backend.new_chatbot
    """
    try:
        data = request.get_json()
        message = data.get('message', '')
        conversation_id = data.get('conversation_id')
        
        if not message:
            return jsonify({
                'success': False,
                'response': 'Message cannot be empty',
                'error': True
            }), 400
        
        # Import the new chatbot v2 (consolidated prompts, faster)
        from backend.new_chatbot import chatbot_v2
        
        # Call the new chatbot (synchronous)
        result = chatbot_v2.chat(message, session_id=conversation_id)
        
        # Format response for legacy API compatibility
        return jsonify({
            'success': result.success,
            'response': result.message or result.error,
            'conversation_id': result.session_id,
            'error': not result.success
        })
        
    except Exception as e:
        logger.error(f"Chat error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'response': 'Sorry, an error occurred. Please try again.',
            'error': True
        }), 500

# Configure Flask-Mail
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER')
app.config['MAIL_PORT'] = os.getenv('MAIL_PORT') # Convert to integer
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS') == 'True'  # Convert string to boolean
app.config['MAIL_USE_SSL'] = os.getenv('MAIL_USE_SSL') == 'True'  # Convert string to boolean
mail = Mail(app)

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('SQLALCHEMY_DATABASE_URI')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = os.getenv('SQLALCHEMY_TRACK_MODIFICATIONS') == 'False'  # Convert string to boolean
db = SQLAlchemy(app)

#=======================================================================================
# WHATSAPP NOTIFICATION SERVICE (DICE API)
#=======================================================================================
class WhatsAppNotificationService:
    _token_cache = {
        "token": None,
        "expires_at": None
    }

    def __init__(self):
        """Initialize WhatsApp service with DICE API config"""
        self.config = {
            "api_username": os.getenv('DICE_API_USERNAME'),
            "api_password": os.getenv('DICE_API_PASSWORD'),
            "auth_url": os.getenv('DICE_AUTH_URL'),
            "whatsapp_url": os.getenv('DICE_WHATSAPP_URL'),
            "template_id": os.getenv('DICE_WHATSAPP_TEMPLATE_ID')
        }
        self.enabled = os.getenv('WHATSAPP_ENABLED', 'False') == 'True'

    def _get_token(self) -> str | None:
        """Get cached OAuth token from DICE API"""
        
        cache = self._token_cache

        # Return cached token if still valid
        if cache["token"] and cache["expires_at"] and cache["expires_at"] > datetime.now(timezone.utc):
            return cache["token"]

        # Generate new token
        try:
            credentials = f"{self.config['api_username']}:{self.config['api_password']}"
            auth_header = base64.b64encode(credentials.encode()).decode()

            headers = {
                "Authorization": f"Basic {auth_header}",
                "User-Agent": "Python Requests"
            }

            response = requests.get(self.config["auth_url"], headers=headers, timeout=10)

            if not response.ok:
                logger.error(f"DICE token generation failed: {response.text}")
                return None

            token = (
                response.json()
                .get("access_token", {})
                .get("data", {})
                .get("access_token")
            )

            if token:
                cache["token"] = token
                cache["expires_at"] = datetime.now(timezone.utc) + timedelta(days=6)
                logger.info("DICE API token generated successfully")

            return token

        except Exception as e:
            logger.error(f"Error getting DICE token: {e}")
            return None

    def send_invoice_cleared_notification(self, vendor_name: str, invoice_number: str, 
                                         invoice_cleared_date: str, mobile_no: str) -> bool:
        """
        Send WhatsApp notification to vendor when invoice is cleared
        
        Args:
            vendor_name: Name of the vendor
            invoice_number: Invoice number
            invoice_cleared_date: Date when invoice was cleared (format: DD-MM-YYYY)
            mobile_no: Vendor's mobile number (10 digits)
        
        Returns:
            bool: True if sent successfully, False otherwise
        """
        
        # Check if WhatsApp is enabled
        if not self.enabled:
            logger.info("WhatsApp notifications are disabled")
            return False

        # Validate inputs
        if not mobile_no or mobile_no.strip() == '':
            logger.warning(f"No mobile number for vendor {vendor_name} - WhatsApp notification skipped")
            return False

        # Get authentication token
        token = self._get_token()
        if not token:
            logger.error("Failed to get DICE API token - WhatsApp notification failed")
            return 
        
        tushar_mobile_number = 9136736171
        try:
            # Clean mobile number (remove spaces, dashes, etc.)
            clean_mobile = ''.join(filter(str.isdigit, str(tushar_mobile_number)))
            
            # Ensure 10 digits for Indian numbers
            if len(clean_mobile) == 10:
                formatted_mobile = f"91{clean_mobile}"
            elif len(clean_mobile) == 12 and clean_mobile.startswith('91'):
                formatted_mobile = f"+{clean_mobile}"
            else:
                formatted_mobile = f"+91{clean_mobile[-10:]}"  # Take last 10 digits

            # Prepare WhatsApp payload (similar to SMS structure)
            payload = {
                "mobile_no": formatted_mobile,
                "channel": "whatsapp",
                "source": "invoice_system",
                "type": "transactional",
                "template_id": self.config['template_id'],
                "template_attr": {
                    "header_value": {
                        "value": vendor_name
                    },
                    "body_value": [
                        invoice_number,
                        invoice_cleared_date,
                    ]
                }
            }

            headers = {
                "Authorization": f"Bearer {token}",
                "User-Agent": "Python Requests",
                "Content-Type": "application/json"
            }

            # Send WhatsApp message
            response = requests.post(
                self.config["whatsapp_url"],
                json=payload,
                headers=headers,
                timeout=10
            )

            # Log the API call
            logger.info(f"WhatsApp API Response for {vendor_name}", extra={
                "status": response.status_code,
                "invoice_number": invoice_number,
                "mobile": formatted_mobile,
                "response": response.text
            })

            # Log activity
            log_activity(f"WhatsApp notification sent to {vendor_name} ({formatted_mobile}) for invoice {invoice_number}")

            # Check if successful
            if response.ok:
                return True

            try:
                return response.json().get("status") in ("success", "SUCCESS")
            except ValueError:
                return False

        except Exception as e:
            logger.error(f"Error sending WhatsApp to {vendor_name}: {e}")
            return False


# Initialize WhatsApp service
whatsapp_service = WhatsAppNotificationService()

#=======================================================================================
# EMAIL NOTIFICATION SERVICE (DICE API)
#=======================================================================================
class EmailNotificationService:

    _token_cache = WhatsAppNotificationService._token_cache  # reuse same token

    DICE_EMAIL_URL = "https://apimartech.auxilo.com/send-message/v1"

    def __init__(self):
        self.config = {
            "api_username": os.getenv('DICE_API_USERNAME'),
            "api_password": os.getenv('DICE_API_PASSWORD'),
            "auth_url": os.getenv('DICE_AUTH_URL'),
        }
        self.enabled = os.getenv('EMAIL_DICE_ENABLED', 'False') == 'True'

    def _get_token(self) -> str | None:
        """Reuse cached DICE token (same logic as WhatsApp service)"""
        cache = self._token_cache
        if cache["token"] and cache["expires_at"] and cache["expires_at"] > datetime.now(timezone.utc):
            return cache["token"]
        try:
            credentials = f"{self.config['api_username']}:{self.config['api_password']}"
            auth_header = base64.b64encode(credentials.encode()).decode()
            headers = {"Authorization": f"Basic {auth_header}", "User-Agent": "Python Requests"}
            response = requests.get(self.config["auth_url"], headers=headers, timeout=10)
            if not response.ok:
                logger.error(f"DICE token generation failed: {response.text}")
                return None
            token = response.json().get("access_token", {}).get("data", {}).get("access_token")
            if token:
                cache["token"] = token
                cache["expires_at"] = datetime.now(timezone.utc) + timedelta(days=6)
            return token
        except Exception as e:
            logger.error(f"Error getting DICE token: {e}")
            return None

    def _send(self, email: str, template_id: str, subject: str, template_attr: dict) -> bool:
        """Core method — builds payload and POSTs to DICE email endpoint"""
        if not self.enabled:
            logger.info("DICE email notifications are disabled")
            return False
        if not email:
            logger.warning("No email address provided — skipping DICE email")
            return False
        token = self._get_token()
        if not token:
            logger.error("Failed to get DICE token — email not sent")
            return False
        try:
            payload = {
                "email": email,
                "channel": "email",
                "source": "Invoice System",
                "type": "transactional",
                "template_id": template_id,
                "email_subject": subject,
                "email_from_name": "Invoice System",
                "template_attr": template_attr
            }
            headers = {
                "Authorization": f"Bearer {token}",
                "User-Agent": "Python Requests",
                "Content-Type": "application/json"
            }
            response = requests.post(self.DICE_EMAIL_URL, json=payload, headers=headers, timeout=10)
            logger.info(f"DICE email sent | template={template_id} | to={email} | status={response.status_code} | response={response.text}")

            # If token was rejected, clear cache and retry once with a fresh token
            if response.status_code == 401:
                logger.warning("DICE token rejected (401) — clearing cache and retrying")
                self._token_cache["token"] = None
                self._token_cache["expires_at"] = None
                token = self._get_token()
                if token:
                    headers["Authorization"] = f"Bearer {token}"
                    response = requests.post(self.DICE_EMAIL_URL, json=payload, headers=headers, timeout=10)
                    logger.info(f"DICE email retry | status={response.status_code} | response={response.text}")

            return response.ok
        except Exception as e:
            logger.error(f"Error sending DICE email to {email}: {e}")
            return False

    # ── Public methods (one per email type) ──────────────────────────────────

    def send_otp(self, email: str, otp: str) -> bool:
        return self._send(
            email=email,
            template_id="Send_Otp",
            subject="Your OTP for Secure Verification",
            template_attr={"otp": otp}
        )

    def send_invoice_added(self, email: str, invoice_number: str, vendor: str,
                           invoice_amount, gst, total_amount, invoice_date,
                           date_received, po_number: str, added_by: str) -> bool:
        return self._send(
            email=email,
            template_id="Invoice_Added",          # ← confirm template ID with DICE team
            subject=f"Invoice System: Invoice #{invoice_number} has been added to the system",
            template_attr={
                "invoice_number": invoice_number,
                "vendor_name": vendor,
                "invoice_amount": str(invoice_amount),
                "gst": str(round(gst, 2)),
                "total_amount": str(total_amount),
                "invoice_date": str(invoice_date),
                "date_received": str(date_received),
                "po_number": po_number or "N/A",
                "added_by": added_by
            }
        )

    def send_invoice_cleared(self, email: str, invoice_number: str, vendor: str,
                              total_amount, date_received, invoice_date,
                              invoice_cleared_date: str, cleared_by: str) -> bool:
        return self._send(
            email=email,
            template_id="Invoice_Cleared",        # ← confirm template ID with DICE team
            subject=f"Invoice System: Invoice #{invoice_number} Cleared",
            template_attr={
                "invoice_number": invoice_number,
                "vendor_name": vendor,
                "total_amount": str(total_amount),
                "date_received": str(date_received),
                "invoice_date": str(invoice_date),
                "invoice_cleared_date": str(invoice_cleared_date),
                "cleared_by": cleared_by
            }
        )

    def send_vendor_approved(self, email: str, vendor_name: str, description: str,
                              requested_by: str, approved_by: str) -> bool:
        return self._send(
            email=email,
            template_id="Vendor_Added",        # ← confirm template ID with DICE team
            subject="Invoice System: New Vendor Addition",
            template_attr={
                "vendor_name": vendor_name,
                "vendor_description": description,
                "requested_by": requested_by,
                "added_on": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "approved_by": approved_by
            }
        )


# Initialize Email service
email_service = EmailNotificationService()
#=======================================================================================
# CHATBOT ENDPOINT - Using new hybrid chatbot  
#=======================================================================================
@app.route('/api/chat/v2', methods=['POST'])
@login_required
def chat_v2_endpoint():
    """
    Enhanced chatbot endpoint with hybrid disambiguation
    Uses the new chatbot from backend.new_chatbot
    """
    try:
        data = request.get_json()
        message = data.get('message', '')
        conversation_id = data.get('conversation_id')
        
        if not message:
            return jsonify({
                'success': False,
                'response': 'Message cannot be empty',
                'error': True
            }), 400
        
        # Import the new chatbot v2 (consolidated prompts, faster)
        from backend.new_chatbot import chatbot_v2
        
        # Call the new chatbot (synchronous, no asyncio needed)
        result = chatbot_v2.chat(message, session_id=conversation_id)
        
        # Handle clarification requests
        if result.needs_clarification:
            return jsonify({
                'success': True,
                'needs_clarification': True,
                'clarification_type': 'entity_selection',
                'message': result.clarifying_question,
                'options': result.options,
                'conversation_id': result.session_id,
                'response': result.message  # For frontend compatibility
            })
        
        # Return the response
        return jsonify({
            'success': result.success,
            'response': result.message or result.error,
            'conversation_id': result.session_id,
            'error': not result.success,
            'sql_query': result.sql_query  # Include SQL for debugging
        })
        
    except Exception as e:
        logger.error(f"Chat V2 error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'response': 'Sorry, an error occurred. Please try again.',
            'error': True,
            'conversation_id': conversation_id if 'conversation_id' in dir() else 'unknown'
        }), 500

# Function to get a MySQL database connection
def get_db_connection():
    """Get connection from pool or create new one"""
    try:
        if db_pool:
            return db_pool.get_connection()
        else:
            # Fallback to direct connection if pool failed
            return mysql.connector.connect(**db_config)
    except Exception as e:
        logger.error(f"Connection error: {e}")
        # Fallback
        return mysql.connector.connect(**db_config)
def generate_po_number(vendor_name,po_date):
    """
    Generate PO number in format: FY25-26/(vendor_shortform)-(today_date)/(increment)
    Example: FY25-26/GOOGLE-13012026/1
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Get vendor shortform
        cursor.execute(
            "SELECT shortforms_of_vendors FROM invoice_uat_db.vendors WHERE vendor_name = %s",
            (vendor_name,)
        )
        vendor = cursor.fetchone()
        
        if not vendor or not vendor['shortforms_of_vendors']:
            return None
        
        shortform = vendor['shortforms_of_vendors'].upper()
        
        # Get today's date in DDMMYYYY format
        selected_date = datetime.strptime(po_date, "%Y-%m-%d").date()
        date_str = selected_date.strftime('%d%m%Y')

        
        # Get current financial year
        current_year = selected_date.year
        if selected_date.month >= 4:  # April to March
            fy_start = current_year
            fy_end = current_year + 1
        else:
            fy_start = current_year - 1
            fy_end = current_year
        
        fy_prefix = f"FY{str(fy_start)[2:]}-{str(fy_end)[2:]}"
        
        # Check existing POs with same vendor and date
        po_prefix = f"{fy_prefix}/{shortform}-{date_str}/"
        
        cursor.execute("""
            SELECT po_number 
            FROM invoice_uat_db.purchase_orders 
            WHERE po_number LIKE %s
            ORDER BY po_number DESC
            LIMIT 1
        """, (f"{po_prefix}%",))
        
        last_po = cursor.fetchone()
        
        if last_po:
            # Extract the increment number
            last_number = int(last_po['po_number'].split('/')[-1])
            increment = last_number + 1
        else:
            increment = 1
        
        po_number = f"{po_prefix}{increment}"
        
        conn.close()
        return po_number
        
    except Exception as e:
        print(f"Error generating PO number: {e}")
        conn.close()
        return None
    
@app.route('/po/generate_number', methods=['POST'])
@login_required
def generate_po_number_api():
    """Generate PO number for selected vendor"""
    try:
        data = request.get_json()
        vendor_name = data.get('vendor_name')
        po_date = data.get('po_date')
        
        if not vendor_name:
            return jsonify({'success': False, 'message': 'Vendor name required'}), 400
        
        po_number = generate_po_number(vendor_name,po_date)
        
        if not po_number:
            return jsonify({'success': False, 'message': 'Could not generate PO number'}), 400
        
        return jsonify({'success': True, 'po_number': po_number})
        
    except Exception as e:
        print(f"Error in generate_po_number_api: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

import hashlib
def hash_otp(otp):
    return hashlib.sha256(otp.encode()).hexdigest()


def generate_otp():
    """Generate cryptographically secure 6-digit OTP"""
    return ''.join(secrets.choice(string.digits) for _ in range(6))

def format_date_ddmmyyyy(date_str):
    if date_str and isinstance(date_str, str) and '-' in date_str:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            return date_str  # fallback if already in correct format or invalid
    return date_str or ""

def previous_month_range():
    today = date.today()
    first_day_this_month = today.replace(day=1)
    last_day_prev_month = first_day_this_month - timedelta(days=1)
    first_day_prev_month = last_day_prev_month.replace(day=1)
    return first_day_prev_month, last_day_prev_month

# def previous_month_range():   #for testing purpose only
#     return date(2025, 11, 1), date(2025, 11, 30)  # TEST MODE

def get_monthly_summary():
    start_date, end_date = previous_month_range()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            COUNT(*) AS total_invoices,
            SUM(CASE WHEN invoice_cleared = 'Yes' THEN total_amount ELSE 0 END) AS cleared_amount,
            SUM(CASE WHEN invoice_cleared = 'Yes' THEN 1 ELSE 0 END) AS cleared_count,
            SUM(CASE WHEN invoice_cleared = 'No' THEN 1 ELSE 0 END) AS uncleared_count
        FROM invoices
        WHERE invoice_date BETWEEN %s AND %s
    """, (start_date, end_date))

    summary = cursor.fetchone()

    cursor.execute("""
        SELECT invoice_date, vendor, invoice_number, total_amount, invoice_cleared
        FROM invoices
        WHERE invoice_date BETWEEN %s AND %s
        ORDER BY invoice_date
    """, (start_date, end_date))

    invoices = cursor.fetchall()

    conn.close()
    return summary, invoices, start_date, end_date

def sanitize_excel(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value

def create_monthly_excel(summary, invoices, start_date, end_date):
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ------------------ SUMMARY SHEET ------------------
        summary_df = pd.DataFrame([{
            "Month": start_date.strftime("%B %Y"),
            "Total Invoices": summary['total_invoices'],
            "Total Cleared Amount": summary['cleared_amount'] or 0,
            "Cleared Invoices": summary['cleared_count'],
            "Uncleared Invoices": summary['uncleared_count']
        }])

        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        # ------------------ INVOICES SHEET ------------------
        invoices_df = pd.DataFrame(invoices)

        # Rename columns for business-friendly Excel
        invoices_df.rename(columns={
            'vendor': 'Vendor Name',
            'invoice_number': 'Invoice Number',
            'invoice_date': 'Invoice Date',
            'total_amount': 'Amount',
            'invoice_cleared': 'Status'
        }, inplace=True)

        # Reorder columns exactly as needed
        invoices_df = invoices_df[
            ['Vendor Name', 'Invoice Number', 'Invoice Date', 'Amount', 'Status']
        ]

        invoices_df.to_excel(writer, index=False, sheet_name="Invoices")

        # ------------------ FORMATTING ------------------
        ws = writer.sheets['Invoices']
        ws.column_dimensions['A'].width = 28  # Vendor
        ws.column_dimensions['B'].width = 20  # Invoice No
        ws.column_dimensions['C'].width = 15  # Date
        ws.column_dimensions['D'].width = 15  # Amount
        ws.column_dimensions['E'].width = 18  # Status

    output.seek(0)
    return output

def send_monthly_email(excel_file, start_date):
    subject = f"Monthly Invoice Report – {start_date.strftime('%B %Y')}"

    recipients_str = os.getenv(
        'REPORT_EMAIL_RECIPIENTS',
        'mihirtendulkar123@gmail.com',
        'tushar.kadam@auxilo.com',
        'tusharkadam1248@gmail.com',
        # 'abhilash.pillai@auxilo.com',
        # 'hemant.dhivar@auxilo.com'
        )
    recipients = [email.strip() for email in recipients_str.split(',')]
    
    msg = Message(
        subject=subject,
        sender=os.getenv('MAIL_USERNAME'),
        recipients=recipients  # 👈 change
    )

    msg.body = f"""
Hello Team,

Please find attached the Monthly Invoice Report.

Month: {start_date.strftime('%B %Y')}

Regards,
Invoice Automation
    """

    msg.attach(
        filename=f"Monthly_Invoice_Report_{start_date.strftime('%b_%Y')}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=excel_file.read()
    )

    mail.send(msg)

# User model
class Users(db.Model,UserMixin):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)  # Added user_name field
    otp = db.Column(db.String(6), nullable=True)
    otp_created_at = db.Column(db.DateTime, nullable=True)
    otp_attempts = db.Column(db.Integer, default=0)
    role = db.Column(db.String(50), nullable=False, default='user')
    department = db.Column(db.String(100), nullable=False, default='marketing')  # ✅ New column 
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    updated_at = db.Column(db.DateTime, default=db.func.current_timestamp(), onupdate=db.func.current_timestamp())

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Users, int(user_id))

class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_email = db.Column(db.String(255), nullable=False)
    action = db.Column(db.String(255), nullable=False)
    department = db.Column(db.String(100), nullable=False, default='marketing')  # ✅ New column 
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

def get_logged_in_user():
    if not current_user.is_authenticated:
        return None
    return db.session.get(Users, current_user.id)

def superadmin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return abort(401)

        if current_user.role != "superadmin":
            return abort(403)

        return f(*args, **kwargs)
    return decorated

# Function to log user activity
def log_activity(action):
    user = get_logged_in_user()
    if not user:
        return

    user_name = user.name or user.email
    role = user.role or "user"

    # Attach role if missing
    if f"{user_name}(" not in action:
        if user_name in action:
            action = action.replace(user_name, f"{user_name}({role})", 1)
        else:
            action = f"{user_name}({role}) {action}"

    log = ActivityLog(
        user_email=user.email,
        action=action,
        department=user.department
    )

    db.session.add(log)
    db.session.commit()

# Endpoint route to send OTP to valid users
@app.route('/send-otp', methods=['POST'])
@csrf.exempt
@limiter.limit("3 per hour")
def send_otp():
    data = request.json
    email = data.get('email')

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    # Check if the email exists in the database
    user = Users.query.filter_by(email=email).first()
    if not user:
        return jsonify({'error': 'Email not found'}), 404

    otp = generate_otp()

    # Update the OTP in the database
    user.otp = hash_otp(otp)
    user.otp_created_at = datetime.utcnow()
    user.otp_attempts = 0
    db.session.commit()
    # Only for me because i doesn't receive emails form DICE api
    print(f"OTP: {otp}")

    # Send the OTP email
    try:
        sent = email_service.send_otp(email=email, otp=otp)
        if not sent:
            return jsonify({'error': 'Failed to send OTP. Please try again later.'}), 500
    except Exception as e:
        return jsonify({'error': 'Failed to send OTP. Please try again later.'}), 500

    # Store email in cookies with expiration
    resp = make_response(jsonify({'message': 'OTP sent successfully'}))
    resp.set_cookie("login_email", email, max_age=300, httponly=True, samesite="Strict")
    return resp


# Endpoint to check weather a valid otp is entered or not
OTP_EXPIRY_MINUTES = 5
@app.route('/verify-otp', methods=['POST'])
@csrf.exempt
@limiter.limit("5 per minute")
def verify_otp():
    data = request.json
    email = data.get('email') or request.cookies.get('login_email')
    otp = data.get('otp')

    if not email or not otp:
        return jsonify({'error': 'Email and OTP are required'}), 400

    user = Users.query.filter_by(email=email, is_active=True).first()
    if not user:
        return jsonify({'valid': False}), 400

    # Lock after 5 attempts
    if user.otp_attempts >= 5:
        return jsonify({'valid': False, 'reason': 'locked'}), 403

    # Expiry
    if not user.otp_created_at or datetime.utcnow() > user.otp_created_at + timedelta(minutes=5):
        user.otp = None
        db.session.commit()
        return jsonify({'valid': False, 'reason': 'expired'}), 400

    # Validate OTP
    if not user.otp or user.otp != hash_otp(otp.strip()):
        user.otp_attempts += 1
        db.session.commit()
        return jsonify({'valid': False}), 400

    # SUCCESS
    user.otp = None
    user.otp_attempts = 0
    db.session.commit()

    login_user(user, remember=True)
    log_activity(f"{user.name}({user.role}) logged in")

    return jsonify({'valid': True})

# Login page route
@app.route('/')
def login():
    return render_template('login.html')

# Route to check OTP
@app.route('/otp')
def otp_page():
    # Check if the user is already logged in
    if request.cookies.get('logged_in') == 'true':
        return redirect(url_for('admin_dashboard'))
    return render_template('otp.html')

# Route for the index page ("Displays invoices")
@app.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ✅ Get the logged-in user
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    query = "SELECT * FROM invoices"
    conditions = ["deleted_at is NULL"]
    params = []

    filter_type = request.args.get('filter')

    if request.method == 'POST':
        vendor = request.form.get('vendor', '').lower()
        invoice_date = request.form.get('invoice_date')
        invoice_start_date = request.form.get('invoice_start_date')
        invoice_end_date = request.form.get('invoice_end_date')
        date_submission = request.form.get('date_submission')
        invoice_number = request.form.get('invoice_number')
        po_number = request.form.get('po_number')
        created_by = request.form.get('created_by', '').lower()

        if vendor:
            conditions.append("LOWER(vendor) LIKE %s")
            params.append(f"%{vendor}%")

        if invoice_date:
            conditions.append("invoice_date = %s")
            params.append(invoice_date)
        elif invoice_start_date and invoice_end_date:
            conditions.append("invoice_date BETWEEN %s AND %s")
            params.extend([invoice_start_date, invoice_end_date])
        elif invoice_start_date:
            conditions.append("invoice_date >= %s")
            params.append(invoice_start_date)
        elif invoice_end_date:
            conditions.append("invoice_date <= %s")
            params.append(invoice_end_date)

        if invoice_number:
            conditions.append("invoice_number LIKE %s")
            params.append(f"%{invoice_number}%")

        if po_number:
            conditions.append("po_number = %s")
            params.append(po_number)

        if created_by:
            conditions.append("LOWER(created_by) LIKE %s")
            params.append(f"%{created_by}%")

    # 🔒 Add department restriction for non-superadmin users
    if user_role != 'superadmin':
        conditions.append("department = %s")
        params.append(user_department)

    # 🧾 Filter by invoice cleared status
    if filter_type == 'cleared':
        conditions.append("invoice_cleared = 'Yes'")
    elif filter_type == 'uncleared':
        conditions.append("invoice_cleared = 'No'")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY invoice_date DESC"

    cursor.execute(query, tuple(params))
    invoices = cursor.fetchall()

    #======================================================================
    # Elapsed time
    #======================================================================

    # Calculate days elapsed for each invoice
    from datetime import datetime, date

    for invoice in invoices:
        if invoice['date_received']:
            # Parse date_received (assuming it's a date object or string)
            if isinstance(invoice['date_received'], str):
                received_date = datetime.strptime(invoice['date_received'], '%Y-%m-%d').date()
            else:
                received_date = invoice['date_received']
            
            # Calculate days elapsed
            if invoice['invoice_cleared'] == 'Yes' and invoice['invoice_cleared_date']:
                # For cleared invoices: stop counting at cleared date
                if isinstance(invoice['invoice_cleared_date'], str):
                    end_date = datetime.strptime(invoice['invoice_cleared_date'], '%Y-%m-%d').date()
                else:
                    end_date = invoice['invoice_cleared_date']
            else:
                # For uncleared invoices: count until today
                end_date = date.today()
            
            # Calculate the difference
            days_elapsed = (end_date - received_date).days
            invoice['days_elapsed'] = days_elapsed
        else:
            invoice['days_elapsed'] = 'N/A'

    no_results_message = "No matching records found." if not invoices else None

    # Dropdowns: fetch distinct values but also filtered by department for non-superadmin
    if user_role == 'superadmin':
        cursor.execute("SELECT DISTINCT vendor FROM invoices")
        vendor_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT created_by FROM invoices")
        created_by_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT invoice_date FROM invoices")
        invoice_date_values = cursor.fetchall()
    else:
        cursor.execute("SELECT DISTINCT vendor FROM invoices WHERE department = %s", (user_department,))
        vendor_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT created_by FROM invoices WHERE department = %s", (user_department,))
        created_by_values = cursor.fetchall()

        cursor.execute("SELECT DISTINCT invoice_date FROM invoices WHERE department = %s", (user_department,))
        invoice_date_values = cursor.fetchall()

    conn.close()

    return render_template(
        'index.html',
        invoices=invoices,
        filters=request.form,
        no_results_message=no_results_message,
        vendor_values=vendor_values,
        created_by_values=created_by_values,
        invoice_date_values=invoice_date_values
    )
    pass

import calendar
from calendar import month_name
from datetime import date, datetime


@app.route('/dashboard', methods=['GET', 'POST'])
@app.route('/dashboard/', methods=['GET', 'POST'])
@login_required
def admin_dashboard():

    today = date.today()

    # ---------- FINANCIAL YEAR SELECTION ----------
    fy_param = request.args.get("fy")

    if fy_param:
        start_year = int(fy_param.split("-")[0])
    else:
        start_year = today.year if today.month >= 4 else today.year - 1

    fy_start = date(start_year, 4, 1)
    fy_end = date(start_year + 1, 3, 31)

    selected_fy = f"{start_year}-{start_year+1}"

    financial_years = []
    for y in range(today.year - 5, today.year + 2):
        financial_years.append(f"{y}-{y+1}")

    # ---------- BASE QUERY ----------
    # ---------- LOGGED IN USER ----------
    user = get_logged_in_user()
    user_role = user.role
    user_department = user.department

    # ---------- DB CONNECTION ----------
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ---------- BASE QUERY ----------
    query = "SELECT * FROM invoices WHERE invoice_date BETWEEN %s AND %s AND deleted_at IS NULL "
    params = [fy_start, fy_end]

    if user_role != "superadmin":
        query += " AND department=%s"
        params.append(user_department)

    # ---------- TABLE FILTERS ----------
    if request.method == "POST":
        vendor = request.form.get("vendor", "").strip()
        start_date = request.form.get("invoice_start_date")
        end_date = request.form.get("invoice_end_date")
        invoice_number = request.form.get("invoice_number", "").strip()
        created_by = request.form.get("created_by", "").strip()

        if vendor:
            query += " AND LOWER(vendor) LIKE %s"
            params.append(f"%{vendor.lower()}%")

        if start_date and end_date:
            query += " AND invoice_date BETWEEN %s AND %s"
            params.append(start_date)
            params.append(end_date)

        elif start_date:
            query += " AND invoice_date >= %s"
            params.append(start_date)

        elif end_date:
            query += " AND invoice_date <= %s"
            params.append(end_date)

        if invoice_number:
            query += " AND invoice_number LIKE %s"
            params.append(f"%{invoice_number}%")

        if created_by:
            query += " AND LOWER(created_by) LIKE %s"
            params.append(f"%{created_by.lower()}%")

    query += " ORDER BY invoice_date DESC"

    cursor.execute(query, tuple(params))
    invoices = cursor.fetchall()

    # ---------- TOTAL INVOICES ----------
    fy_params = [fy_start, fy_end]
    if user_role != "superadmin":
        fy_params.append(user_department)

    # cursor.execute("""
    #     SELECT COUNT(*) total FROM invoices
    #     WHERE invoice_date BETWEEN %s AND %s 
    #     AND deleted_at IS NULL
    #     """ + (" AND department=%s" if user_role != "superadmin" else ""),
    #     tuple(fy_params)
    # )
    # total_invoices = cursor.fetchone()['total']


    # ---------- OVERALL POOL ----------
    
    cursor.execute("""
        SELECT SUM(total_amount) total FROM invoices
        WHERE invoice_cleared='Yes' AND invoice_date BETWEEN %s AND %s
        AND deleted_at IS NULL
        """ + (" AND department=%s" if user_role != "superadmin" else ""),
        tuple(fy_params)
    )
    overall_pool = cursor.fetchone()['total'] or 0

    # ---------- CURRENT MONTH SPEND ----------
    current_month = today.month
    current_year = today.year

    cursor.execute("""
        SELECT SUM(total_amount) AS total
        FROM invoices
        WHERE invoice_cleared = 'Yes'
        AND MONTH(invoice_date) = %s
        AND YEAR(invoice_date) = %s
        AND deleted_at IS NULL
    """ + (" AND department=%s" if user_role != "superadmin" else ""),
    tuple(
        [current_month, current_year] +
        ([user_department] if user_role != "superadmin" else [])
    ))

    monthly_pool = cursor.fetchone()['total'] or 0


    # ---------- MONTHLY CHART ----------
    cursor.execute("""
        SELECT MONTH(invoice_date) m, SUM(total_amount) total
        FROM invoices
        WHERE invoice_cleared='Yes'
        AND invoice_date BETWEEN %s AND %s
        AND deleted_at IS NULL
        """ + (" AND department=%s" if user_role != "superadmin" else "") + """
        GROUP BY MONTH(invoice_date)
        ORDER BY MONTH(invoice_date)
    """, tuple(fy_params))

    rows = cursor.fetchall()

    monthly_labels = []
    monthly_values = []

    for m in [4,5,6,7,8,9,10,11,12,1,2,3]:
        monthly_labels.append(calendar.month_name[m])
        value = next((r['total'] for r in rows if r['m'] == m), 0)
        monthly_values.append(float(value or 0))

    # ---------- PENDING ----------
    total_pending_invoices = sum(1 for i in invoices if i['invoice_cleared'] != 'Yes' and i['deleted_at'] is None)

    delayed_30 = sum(1 for i in invoices if i['invoice_cleared']=='No'
                     and i['invoice_date'] and (today - i['invoice_date']).days > 30 and i['deleted_at'] is None)

    delayed_60 = sum(1 for i in invoices if i['invoice_cleared']=='No'
                     and i['invoice_date'] and (today - i['invoice_date']).days > 60 and i['deleted_at'] is None)

    # ---------- TOP 3 VENDORS ----------
    cursor.execute("""
        SELECT vendor, SUM(total_amount) total
        FROM invoices
        WHERE invoice_cleared='Yes'
        AND invoice_date BETWEEN %s AND %s
        AND deleted_at IS NULL
        """ + (" AND department=%s" if user_role != "superadmin" else "") + """
        GROUP BY vendor
        ORDER BY total DESC
        LIMIT 3
    """, tuple(fy_params))
    top_vendors = cursor.fetchall()

    # ---------- TOP 3 TAGS ----------
    cursor.execute("""
        SELECT tag1 AS tag, SUM(total_amount) total
        FROM invoices
        WHERE tag1 IS NOT NULL AND tag1!=''
        AND invoice_cleared='Yes'
        AND invoice_date BETWEEN %s AND %s
        AND deleted_at IS NULL
        """ + (" AND department=%s" if user_role != "superadmin" else "") + """
        GROUP BY tag1
        ORDER BY total DESC
        LIMIT 3
    """, tuple(fy_params))
    top_tags = cursor.fetchall()

    # ---------- VENDOR LIST ----------
    cursor.execute("SELECT DISTINCT vendor FROM invoices WHERE vendor!='' AND vendor IS NOT NULL ORDER BY vendor")
    vendor_list = [r['vendor'] for r in cursor.fetchall()]
    
    # ---------- VENDOR TRENDS DATA ----------
    vendor_trends_data = get_vendor_monthly_trends(selected_fy)

    # ---------- TAG LIST ----------
    cursor.execute("SELECT DISTINCT tag1 FROM invoices WHERE tag1!='' AND tag1 IS NOT NULL")
    tag1_list = [r['tag1'] for r in cursor.fetchall()]

    # ---------- PARTICULAR MONTH SPENDING ----------
    selected_month = request.args.get("month", type=int)
    selected_year = request.args.get("year", type=int)

    # Default: current month of FY
    if not selected_month:
        selected_month = today.month

    if not selected_year:
        # if month is Jan/Feb/Mar → belongs to next year of FY
        if selected_month in [1,2,3]:
            selected_year = start_year + 1
        else:
            selected_year = start_year

    month_query = """
        SELECT SUM(total_amount) AS total
        FROM invoices
        WHERE invoice_cleared='Yes'
        AND MONTH(invoice_date)=%s
        AND YEAR(invoice_date)=%s
    """

    month_params = [selected_month, selected_year]

    if user_role != "superadmin":
        month_query += " AND department=%s"
        month_params.append(user_department)

    cursor.execute(month_query, tuple(month_params))
    month_spending = cursor.fetchone()['total'] or 0

    # ---------- FY PARAMS FOR ALL DASHBOARD KPI QUERIES ----------
    fy_params = [fy_start, fy_end]
    if user_role != "superadmin":
        fy_params.append(user_department)


    # ---------- TOTAL INVOICES ----------
    cursor.execute("""
        SELECT COUNT(*) total 
        FROM invoices
        WHERE invoice_date BETWEEN %s AND %s
        AND deleted_at IS NULL
        """ + (" AND department=%s" if user_role != "superadmin" else ""),
        tuple(fy_params)
    )
    total_invoices = cursor.fetchone()['total']

    # ---------- TAG1 MONTHLY TRENDS ----------
    trend_tag = request.args.get('trend_tag', None)
    tag1_trends_data = get_tag1_monthly_trends(selected_fy, trend_tag)

    conn.close()

    return render_template(
        'admin_dashboard.html',

        invoices=invoices,
        total_invoices=total_invoices,
        total_cleared_invoices=sum(1 for i in invoices if i['invoice_cleared']=='Yes'),

        overall_pool=overall_pool,
        monthly_pool=monthly_pool,

        today=today,

        delayed_30=delayed_30,
        delayed_60=delayed_60,
        total_pending_invoices=total_pending_invoices,

        top_vendors=top_vendors,
        top_tags=top_tags,

        financial_years=financial_years,
        selected_fy=selected_fy,

        monthly_labels=monthly_labels,
        monthly_values=monthly_values,

        calendar=calendar,

        tag1_list=tag1_list,
        tag1_trends_data=tag1_trends_data,
        trend_tag=trend_tag,

        vendor_list = vendor_list,
        vendor_trends_data = vendor_trends_data,

    )

@app.route("/api/tag1_trends")
@login_required
def api_tag1_trends():
    """
    API endpoint to get Tag1 monthly trends data via AJAX
    Returns JSON with labels and data for Chart.js
    """
    fy_param = request.args.get("fy")
    trend_tag = request.args.get("trend_tag", None)
    
    if not fy_param:
        return jsonify({"error": "Financial year parameter required"}), 400
    
    # Get trends data
    trends_data = get_tag1_monthly_trends(fy_param, trend_tag)
    
    return jsonify(trends_data)

@app.route('/api/vendor_trends')
@login_required
def api_vendor_trends():
    """API endpoint for vendor spending trends"""
    try:
        selected_fy = request.args.get('fy', '2024-2025')
        trend_vendor = request.args.get('trend_vendor', None)
        
        # If empty string, treat as None
        if trend_vendor == '':
            trend_vendor = None
            
        trends_data = get_vendor_monthly_trends(selected_fy, trend_vendor)
        return jsonify(trends_data)
    except Exception as e:
        logger.error(f"Error in vendor trends API: {e}")
        return jsonify({'error': str(e)}), 500
    
@app.route("/api/month_spend")
@login_required
def api_month_spend():
    month = int(request.args.get("month"))
    year = int(request.args.get("year"))

    user = get_logged_in_user()
    user_role = user.role
    user_department = user.department

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT SUM(total_amount) AS total
        FROM invoices
        WHERE MONTH(invoice_date)=%s 
        AND YEAR(invoice_date)=%s
    """
    params = [month, year]

    if user_role != "superadmin":
        query += " AND department=%s"
        params.append(user_department)

    cursor.execute(query, tuple(params))
    value = cursor.fetchone()['total'] or 0

    conn.close()

    return jsonify({"amount": float(value)})

@app.route("/api/top_criteria")
@login_required
def api_top_criteria():
    tag = request.args.get("tag")
    from_month = int(request.args.get("from_month"))
    to_month = int(request.args.get("to_month"))

    user = get_logged_in_user()
    user_role = user.role
    user_department = user.department

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT SUM(total_amount) AS total
        FROM invoices
        WHERE tag1=%s
        AND MONTH(invoice_date) BETWEEN %s AND %s
        AND invoice_cleared='Yes'
    """
    params = [tag, from_month, to_month]

    if user_role != "superadmin":
        query += " AND department=%s"
        params.append(user_department)

    cursor.execute(query, tuple(params))
    value = cursor.fetchone()['total'] or 0

    conn.close()

    return jsonify({"amount": float(value)})
@app.route("/api/invoices", methods=["GET", "POST"])
@login_required
def get_invoices_api():

    draw = int(request.values.get("draw", 1))
    start = int(request.values.get("start", 0))
    length = int(request.values.get("length", 10))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Base Query
    base_query = "FROM invoices WHERE 1=1"
    params = []

    user = get_logged_in_user()
    if user.role != "superadmin":
        base_query += " AND department=%s"
        params.append(user.department)

    # -------- FILTERS ----------
    vendor = request.values.get("vendor", "").strip().lower()
    if vendor:
        base_query += " AND LOWER(vendor) LIKE %s"
        params.append(f"%{vendor}%")

    start_date = request.values.get("invoice_start_date")
    end_date = request.values.get("invoice_end_date")

    if start_date and end_date:
        base_query += " AND invoice_date BETWEEN %s AND %s"
        params.extend([start_date, end_date])
    elif start_date:
        base_query += " AND invoice_date >= %s"
        params.append(start_date)
    elif end_date:
        base_query += " AND invoice_date <= %s"
        params.append(end_date)

    invoice_number = request.values.get("invoice_number", "")
    if invoice_number:
        base_query += " AND invoice_number LIKE %s"
        params.append(f"%{invoice_number}%")

    created_by = request.values.get("created_by", "").strip().lower()
    if created_by:
        base_query += " AND LOWER(created_by) LIKE %s"
        params.append(f"%{created_by}%")

    # -------- TOTAL RECORDS ----------
    cursor.execute(f"SELECT COUNT(*) total {base_query}", tuple(params))
    recordsFiltered = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) total FROM invoices")
    recordsTotal = cursor.fetchone()["total"]

    # -------- PAGINATED RESULT ----------
    cursor.execute(f"""
        SELECT *
        {base_query}
        ORDER BY invoice_date DESC
        LIMIT %s OFFSET %s
    """, tuple(params + [length, start]))

    data = cursor.fetchall()
    conn.close()

    return jsonify({
        "draw": draw,
        "recordsTotal": recordsTotal,
        "recordsFiltered": recordsFiltered,
        "data": data
    })

# Route for download excel button("With / Without filters")
@app.route('/download_excel', methods=['POST'])
def download_excel():
    vendor = request.form.get('vendor')
    invoice_date = request.form.get('invoice_date')
    date_submission = request.form.get('date_submission')
    invoice_number = request.form.get('invoice_number')
    po_number = request.form.get('po_number')
    created_by = request.form.get('created_by')

    # Building the raw SQL query
    query = "SELECT * FROM invoices"
    conditions = []
    filters = {}

    # Apply filters based on form input
    if vendor:
        conditions.append("vendor = %s")
        filters['vendor'] = vendor
    if invoice_date:
        conditions.append("invoice_date = %s")
        filters['invoice_date'] = invoice_date
    if date_submission:
        conditions.append("date_submission = %s")
        filters['date_submission'] = date_submission
    if invoice_number:
        conditions.append("invoice_number = %s")
        filters['invoice_number'] = invoice_number
    if po_number:
        conditions.append("po_number = %s")
        filters['po_number'] = po_number
    if created_by:
        conditions.append("created_by = %s")
        filters['created_by'] = created_by

    # Append conditions to the query
    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # Execute the SQL query
    conn = get_db_connection()  # Replace this with your actual connection function
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Convert to DataFrame for easy export to Excel
    df = pd.DataFrame([{
        'Invoice Date': invoice['invoice_date'],
        'Date Received': invoice['date_received'],
        'Vendor': invoice['vendor'],
        'Invoice Number': invoice['invoice_number'],
        'PO Number': invoice['po_number'],
        'MSME': invoice['msme'],
        'Invoice Amount': invoice['invoice_amount'],
        'GST': invoice['gst'],
        'Total Amount': invoice['total_amount'],
        'Date of Submission': invoice['date_submission'],
        'Approved By': invoice['approved_by'],
        'HOD Approval': invoice['hod_values'],
        'CEO Approval': invoice['ceo_values'],
        'Reviewed By': invoice['reviewed_by'],
        'Created By': invoice['created_by'],
        'Tag1': invoice['tag1'],
        'Tag2': invoice['tag2'],
        'Invoice Cleared': invoice['invoice_cleared'],
        'Cleared Date': invoice['invoice_cleared_date'],
    } for invoice in invoices])

    # Create a bytes buffer for the Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Seek to the beginning of the stream
    output.seek(0)

    # Send the Excel file to the user for download
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='filtered_invoices.xlsx')


    # Execute the SQL query
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(filters.values()))
    invoices = cursor.fetchall()

    # Check if no invoices are returned for the applied filters
    no_results_message = None
    if not invoices:
        no_results_message = "No matching records found."

    # Total number of invoices
    cursor.execute('SELECT COUNT(*) FROM invoices')
    total_invoices = cursor.fetchone()['COUNT(*)']

    # Total cleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "Yes"')
    total_cleared_invoices = cursor.fetchone()['COUNT(*)']

    # Total uncleared invoices
    cursor.execute('SELECT COUNT(*) FROM invoices WHERE invoice_cleared = "No"')
    total_uncleared_invoices = cursor.fetchone()['COUNT(*)']

    conn.close()

    return render_template(
        'admin_dashboard.html', invoices=invoices,
        total_invoices=total_invoices,
        total_cleared_invoices=total_cleared_invoices,
        total_uncleared_invoices=total_uncleared_invoices,
        no_results_message=no_results_message  # Pass message to template
    )
    pass

@app.route('/add', methods=('GET', 'POST'))
@login_required
def add_invoice():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    #  Get logged-in user info
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    #  Fetch department-specific vendors
    cursor.execute('SELECT * FROM vendors WHERE department = %s AND deleted_at IS NULL', (user_department,))
    vendors = cursor.fetchall()

    #  Department-specific dropdown helper
    def get_dropdown_values(value_type):
        cursor.execute(
            "SELECT value FROM dropdown_values WHERE type = %s AND department = %s AND is_active = TRUE",
            (value_type, user_department)
        )
        return [row['value'] for row in cursor.fetchall()]
        
    # Fetch all dropdown values
    approved_by_values = get_dropdown_values('approved_by')
    created_by_values = get_dropdown_values('created_by')
    hod_values = get_dropdown_values('hod')
    ceo_values = get_dropdown_values('ceo')
    reviewed_by_values = get_dropdown_values('reviewed_by')
    tag1 = get_dropdown_values('tag1')
    tag2 = get_dropdown_values('tag2')

    if request.method == 'POST':
        # Fetch form data
        invoice_date = request.form['invoice_date']
        date_received = request.form['date_received']
        vendor = request.form['vendor']
        mobile_no = request.form['mobile_no']
        invoice_number = request.form['invoice_number']
        date_submission = request.form['date_submission']
        approved_by = request.form.get('approved_by')
        created_by = user.name
        po_approved = request.form['po_approved']
        po_number = request.form['po_number']
        agreement_signed = request.form['agreement_signed']
        po_expiry_date = request.form.get('po_expiry_date') or None
        agreement_signed_date = request.form.get('agreement_signed_date') or None
        hod_approval = request.form.get('hod_values')
        ceo_approval = request.form.get('ceo_values')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')
        invoice_amount = float(request.form['invoice_amount'])
        gst = invoice_amount * 0.18
        total_amount = invoice_amount + gst
        total_amount = round(total_amount, 2)
        isd = request.form.get('isd', 'No')  # Default to 'No' if not selected
        invoice_cleared = request.form.get('invoice_cleared', 'No')  # Default to 'No'
        invoice_cleared_date = request.form.get('invoice_cleared_date') or None
        
        msme = request.form['msme']

        total_amount_words = num2words(total_amount, to='currency', currency='INR', lang='en_IN').upper().replace(",", "")

        # Validate form inputs
        if not invoice_date or not date_received or not vendor or not invoice_number or not date_submission or not created_by or not invoice_amount:
            flash('All fields are required!')
        else:
            # Check if the invoice number already exists
            cursor.execute('SELECT * FROM invoices WHERE invoice_number = %s', (invoice_number,))
            existing_invoice = cursor.fetchone()

            if existing_invoice:
                flash('Invoice number already exists. Please enter a unique invoice number.')
                conn.close()

                form_data = {
                    'invoice_date': invoice_date,
                    'date_received': date_received,
                    'vendor': vendor,
                    'mobile_no': mobile_no,
                    'invoice_number': invoice_number,
                    'date_submission': date_submission,
                    'approved_by': approved_by,
                    'po_approved': po_approved,
                    'po_number': po_number,
                    'agreement_signed': agreement_signed,
                    'po_expiry_date': po_expiry_date,
                    'agreement_signed_date': agreement_signed_date,
                    'hod_values': hod_approval,
                    'ceo_values': ceo_approval,
                    'reviewed_by': reviewed_by,
                    'tag1': tag1,
                    'tag2': tag2,
                    'invoice_amount': invoice_amount,
                    'isd': isd,
                    'msme': msme,
                    'invoice_cleared': invoice_cleared,
                    'invoice_cleared_date': invoice_cleared_date
                }
                return render_template(
                    'add_invoice.html',
                    vendors=vendors,
                    approved_by_values=approved_by_values,
                    created_by_values=created_by_values,
                    hod_values=hod_values,
                    ceo_values=ceo_values,
                    reviewed_by_values=reviewed_by_values,
                    tag1=tag1,
                    tag2=tag2,
                    form_data=form_data  
                )
            else:
                try:
                    # Start transaction
                    conn.autocommit = False
                    
                    cursor.execute(
                        '''INSERT INTO invoices (
                            invoice_date, date_received, vendor, mobile_no, invoice_number, po_approved, po_number, po_expiry_date,
                            agreement_signed, agreement_signed_date, date_submission, approved_by, created_by, tag1, tag2,
                            invoice_amount, gst, total_amount, isd, msme, hod_values, ceo_values, reviewed_by,
                            invoice_cleared, invoice_cleared_date,department
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                        (
                            invoice_date, date_received, vendor, mobile_no, invoice_number, po_approved, po_number, po_expiry_date,
                            agreement_signed, agreement_signed_date, date_submission, approved_by, created_by, tag1, tag2,
                            invoice_amount, gst, total_amount, isd, msme, hod_approval, ceo_approval, reviewed_by,
                            invoice_cleared, invoice_cleared_date,user_department
                        )
                    )
                    
                    # Commit transaction
                    conn.commit()
                    conn.autocommit = True
                    
                    flash('Invoice added successfully!')

                    #----------------------------------------------------------------
                    # Sending email notification when invoice is added
                    #----------------------------------------------------------------
                    try:
                        invoice_added_recipients = os.getenv('INVOICE_ADDED_RECIPIENTS', '')
                        if invoice_added_recipients:
                            for recipient in [r.strip() for r in invoice_added_recipients.split(',') if r.strip()]:
                                email_service.send_invoice_added(
                                    email=recipient,
                                    invoice_number=invoice_number,
                                    vendor=vendor,
                                    invoice_amount=invoice_amount,
                                    gst=gst,
                                    total_amount=total_amount,
                                    invoice_date=invoice_date,
                                    date_received=date_received,
                                    po_number=po_number,
                                    added_by=user.name
                                )
                            logger.info(f"DICE invoice added email sent for {invoice_number} to {invoice_added_recipients}")
                    except Exception as e:
                        logger.error(f"Failed to send invoice added email for {invoice_number}: {e}")
                    
                except mysql.connector.IntegrityError as e:
                    # Rollback on duplicate or constraint error
                    conn.rollback()
                    conn.autocommit = True
                    
                    if e.errno == 1062:  # Duplicate entry
                        flash('Invoice number already exists. Please use a different number.')
                        conn.close()
                        form_data = {
                            'invoice_date': invoice_date,
                            'date_received': date_received,
                            'vendor': vendor,
                            'mobile_no': mobile_no,
                            'invoice_number': invoice_number,
                            'date_submission': date_submission,
                            'approved_by': approved_by,
                            'po_approved': po_approved,
                            'po_number': po_number,
                            'agreement_signed': agreement_signed,
                            'po_expiry_date': po_expiry_date,
                            'agreement_signed_date': agreement_signed_date,
                            'hod_values': hod_approval,
                            'ceo_values': ceo_approval,
                            'reviewed_by': reviewed_by,
                            'tag1': tag1,
                            'tag2': tag2,
                            'invoice_amount': invoice_amount,
                            'isd': isd,
                            'msme': msme,
                            'invoice_cleared': invoice_cleared,
                            'invoice_cleared_date': invoice_cleared_date
                        }
                        return render_template(
                        'add_invoice.html',
                        vendors=vendors,
                        approved_by_values=approved_by_values,
                        created_by_values=created_by_values,
                        hod_values=hod_values,
                        ceo_values=ceo_values,
                        reviewed_by_values=reviewed_by_values,
                        tag1=tag1,
                        tag2=tag2,
                        form_data = form_data
                        )
                    else:
                        flash(f'Database error: {str(e)}')
                        conn.close()
                        return render_template(
                            'add_invoice.html',
                            vendors=vendors,
                            approved_by_values=approved_by_values,
                            created_by_values=created_by_values,
                            hod_values=hod_values,
                            ceo_values=ceo_values,
                            reviewed_by_values=reviewed_by_values,
                            tag1=tag1,
                            tag2=tag2
                        )
                        
                except Exception as e:
                    # Rollback on any other error
                    conn.rollback()
                    conn.autocommit = True
                    flash(f'Failed to add invoice: {str(e)}')
                    conn.close()
                    return render_template(
                        'add_invoice.html',
                        vendors=vendors,
                        approved_by_values=approved_by_values,
                        created_by_values=created_by_values,
                        hod_values=hod_values,
                        ceo_values=ceo_values,
                        reviewed_by_values=reviewed_by_values,
                        tag1=tag1,
                        tag2=tag2
                    )

                actor_role = current_user.role
                user_name = request.cookies.get('name') or request.cookies.get('email')
                log_activity(f"{user_name}({actor_role}) added invoice ({invoice_number}) for vendor {vendor}")

                # Generate Excel file using openpyxl
                #template_path = "static/excel_templates/template.xlsx"
                template_path = "static/excel_templates/Vendor_Payment_form.xlsx"

                wb = load_workbook(template_path)
                ws = wb.active

                invoice_date_fmt = format_date_ddmmyyyy(invoice_date)

                # Populate data into the template
                ws['A7'] = sanitize_excel(vendor)
                ws['E6'] = sanitize_excel(invoice_date_fmt)
                ws['E7'] = sanitize_excel(invoice_number)
                ws['B8'] = sanitize_excel(po_approved)
                ws['B9'] = sanitize_excel(agreement_signed)
                ws['E8'] = sanitize_excel(po_expiry_date)
                ws['E9'] = sanitize_excel(agreement_signed_date)
                ws['B13'] = sanitize_excel(isd)  # Or any cell where ISD info should appear
                ws['E11'] = sanitize_excel(msme)
                ws['E12'] = sanitize_excel(mobile_no)
                ws['F25'] = sanitize_excel(total_amount)
                ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F25'].number_format = numbers.FORMAT_NUMBER_00

                ws['F35'] = total_amount
                ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F35'].number_format = numbers.FORMAT_NUMBER_00

                # ws['B35'] = total_amount_words
                if ws['A35'].value:
                    ws['A35'] = f"{ws['A35'].value} {total_amount_words}"
                else:
                    ws['A35'] = total_amount_words
                ws['A35'].alignment = Alignment(wrap_text=True)
                ws.column_dimensions['A'].width = 50 # Increase the width of column A
                ws.row_dimensions[35].height = 40  # Increase the height of row 35

                if ws['A25'].value:  # Check if there is already a value in the cell
                    ws['A25'] = f"{ws['A25'].value} ({tag1})"
                else:
                    ws['A25'] = f"Marketing Expenses ({tag1})"

                # ws['D11'] = date_submission

                if ws['D47'].value:  # Check if there's already a value in the cell
                    ws['D47'] = f"{ws['D47'].value} {created_by}"
                else:
                    ws['D47'] = sanitize_excel(created_by)

                # For approved_by
                if ws['D41'].value:
                    ws['D41'] = f"{ws['D41'].value} {approved_by}"
                else:
                    ws['D41'] = sanitize_excel(approved_by)

                # For reviewed_by
                if ws['D39'].value:
                    ws['D39'] = f"{ws['D39'].value} {reviewed_by}"
                else:
                    ws['D39'] = sanitize_excel(reviewed_by)

                # For hod_approval
                if ws['D43'].value:
                    ws['D43'] = f"{ws['D43'].value} {hod_approval}"
                else:
                    ws['D43'] = hod_approval

                # For ceo_approval
                if ws['D45'].value:
                    ws['D45'] = f"{ws['D45'].value} {ceo_approval}"
                else:
                    ws['D45'] = ceo_approval

                # Save the populated Excel file to a BytesIO stream
                excel_file = BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                # Return the Excel file as a response
                return send_file(
                    excel_file,
                    as_attachment=True,
                    download_name=f"invoice_{invoice_number}.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    conn.close()

    return render_template(
        'add_invoice.html',
        vendors=vendors,
        approved_by_values=approved_by_values,
        created_by_values=created_by_values,
        hod_values=hod_values,
        ceo_values=ceo_values,
        reviewed_by_values=reviewed_by_values,
        tag1=tag1,
        tag2=tag2
    )
    pass

@app.route('/edit/<int:id>', methods=('GET', 'POST'))
@login_required
def edit_invoice(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get the logged-in user
    user = get_logged_in_user()
    if not user:
        flash("User not found or not logged in. Please login again.")
        return redirect(url_for('login'))

    user_role = user.role
    user_department = user.department

    # Fetch the invoice to be edited
    cursor.execute('SELECT * FROM invoices WHERE id = %s', (id,))
    invoice = cursor.fetchone()

    if not invoice:
        flash("Invoice not found.")
        return redirect(url_for('dashboard'))
    
    if invoice['invoice_cleared'] == 'Yes':
        flash("This invoice is cleared and cannot be edited.", "warning")
        return redirect(url_for('index'))

    # Fetch vendors based on department
    cursor.execute('SELECT * FROM vendors WHERE department = %s', (user_department,))
    vendors = cursor.fetchall()

    # Helper to fetch dropdown values department-wise
    def get_dropdown_values(value_type):
        cursor.execute(
            "SELECT value FROM dropdown_values WHERE type = %s AND department = %s AND is_active = TRUE",
            (value_type, user_department)
        )
        return [row['value'] for row in cursor.fetchall()]

    approved_by_values = get_dropdown_values('approved_by')
    created_by_values = get_dropdown_values('created_by')
    hod_values = get_dropdown_values('hod')
    ceo_values = get_dropdown_values('ceo')
    reviewed_by_values = get_dropdown_values('reviewed_by')
    tag1 = get_dropdown_values('tag1')
    tag2 = get_dropdown_values('tag2')

    if request.method == 'POST':
        invoice_date = request.form['invoice_date']
        date_received = request.form['date_received']
        vendor = request.form['vendor']
        mobile_no = request.form['mobile_no']
        invoice_number = request.form['invoice_number']
        po_approved = request.form['po_approved']
        po_number = request.form['po_number']
        agreement_signed = request.form['agreement_signed']
        date_submission = request.form['date_submission']
        isd = request.form.get('isd')  # Default to 'No' if not selected
        approved_by = request.form.get('approved_by')
        hod_approval = request.form.get('hod_values')
        ceo_approval = request.form.get('ceo_values')
        reviewed_by = request.form.get('reviewed_by')
        tag1 = request.form.get('tag1')
        tag2 = request.form.get('tag2')
        po_expiry_date = request.form.get('po_expiry_date')  # Optional
        agreement_signed_date = request.form.get('agreement_signed_date')  # Optional
        po_expiry_date = po_expiry_date if po_expiry_date else None
        agreement_signed_date = agreement_signed_date if agreement_signed_date else None
        created_by = invoice['created_by']
        invoice_cleared = request.form['invoice_cleared']

        if invoice_cleared == 'Yes':
            invoice_cleared_date = request.form.get('invoice_cleared_date') or date.today()
           
        else:
            invoice_cleared_date = None

        invoice_amount = float(request.form['invoice_amount'])
        gst = invoice_amount * 0.18
        total_amount = invoice_amount + gst
        total_amount_words = num2words(total_amount, to='currency', currency='INR', lang='en_IN')
        total_amount_words = total_amount_words.upper().replace(",", "")
        msme = request.form['msme']

        if not invoice_date or not date_received or not vendor or not invoice_number or not date_submission or not invoice_amount:
            flash('All fields are required!')
        else:
            cursor.execute(
                '''
                UPDATE invoices 
                SET invoice_date = %s, 
                    date_received = %s, 
                    vendor = %s, 
                    mobile_no = %s,
                    invoice_number = %s, 
                    po_approved = %s, 
                    po_number = %s, 
                    po_expiry_date = %s, 
                    agreement_signed = %s, 
                    agreement_signed_date = %s, 
                    date_submission = %s, 
                    approved_by = %s, 
                    created_by = %s, 
                    tag1 = %s, 
                    tag2 = %s, 
                    invoice_amount = %s, 
                    gst = %s,
                    isd = %s, 
                    total_amount = %s, 
                    msme = %s, 
                    hod_values = %s, 
                    ceo_values = %s, 
                    reviewed_by = %s, 
                    invoice_cleared = %s, 
                    invoice_cleared_date = %s
                WHERE id = %s
                ''',
                (
                    invoice_date, date_received, vendor, mobile_no, invoice_number, 
                    po_approved, po_number, po_expiry_date, agreement_signed, 
                    agreement_signed_date, date_submission, approved_by, created_by, 
                    tag1, tag2, invoice_amount, gst,isd ,total_amount, msme, 
                    hod_approval, ceo_approval, reviewed_by, invoice_cleared, 
                    invoice_cleared_date, id
                )
            )
            conn.commit()
            #----------------------------------------------------------------
            # Sending email when invoice is cleared
            #----------------------------------------------------------------
            email_sent = False
            whatsapp_sent = False
            try:
                recipients = os.getenv('VENDOR_ADDED_RECIPIENTS', '').split(',')
                for recipient in [r.strip() for r in recipients if r.strip()]:
                    email_service.send_invoice_cleared(
                        email=recipient,
                        invoice_number=invoice_number,
                        vendor=vendor,
                        total_amount=total_amount,
                        date_received=date_received,
                        invoice_date=invoice['invoice_date'],
                        invoice_cleared_date=invoice_cleared_date,
                        cleared_by=current_user.name
                    )
                email_sent = True
                log_activity(f'Email sent: invoice {invoice_number} for {vendor} cleared')
            except Exception as e:
                logger.error(f"Failed to send invoice cleared email: {e}")
                email_sent = False

            #----------------------------------------------------------------
            # Sending WhatsApp notification to vendor when invoice is cleared
            #----------------------------------------------------------------
            if invoice_cleared == 'Yes':
                try:
                    if isinstance(invoice_cleared_date,date):
                        formatted_date = invoice_cleared_date.strftime('%d-%m-%Y')                

                    elif isinstance(invoice_cleared_date, str):

                        try:
                            date_obj = datetime.strptime(invoice_cleared_date).strftime('%Y-%m-%d')
                            formatted_date = date_obj.strftime('%d-%m-%Y')
                        except:
                            formatted_date = invoice_cleared_date

                    else:
                        formatted_date = datetime.now().strftime('%d-%m-%Y')

                    # Send Whatsapp notification
                    whatsapp_sent = whatsapp_service.send_invoice_cleared_notification(
                        vendor_name=vendor,
                        invoice_number=invoice_number,
                        invoice_cleared_date=formatted_date,
                        mobile_no=mobile_no
                    )

                    if whatsapp_sent:
                        logger.info(f"Whatsapp notification sent successfully to {vendor} for invoice: {invoice_number}")
                        log_activity(f"Whatsapp notification sent successfully to {vendor} for invoice: {invoice_number}")

                    else:
                        logger.warning(f"Whatsapp notification failed for {vendor}, invoice: {invoice_number}")

                except Exception as e:
                    logger.error(f"Failed to send WhatsApp notification to vendor: {e}")
                    whatsapp_sent = False

            message = 'Invoice updated successfully!'
            if invoice_cleared == 'Yes':
                if email_sent and whatsapp_sent:
                    message += ' Email and Whatsapp notifications sent'
                elif email_sent:
                    message += ' Email sent. (Whatsapp notification failed)'
                elif whatsapp_sent:
                    message += ' WhatsApp sent. (Email notification failed)'
                else:
                    message += (' Notifications failed')
            flash(message)
            actor_role = current_user.role
            user_name = request.cookies.get('name') or request.cookies.get('email')
            log_activity(f"{user_name}({actor_role}) edited invoice ({invoice_number}) for vendor {vendor}")

            # Skip Excel generation if invoice is cleared
            if invoice_cleared == 'Yes':
                conn.close()
                return redirect(url_for('index'))  # Redirect to a different view after successful update
            
            # Generate the updated invoice Excel
            # template_path = "static/excel_templates/template.xlsx"
            template_path = "static/excel_templates/Vendor_Payment_form.xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            invoice_date_fmt = format_date_ddmmyyyy(invoice_date)

            # Populate data into the template
            ws['A7'] = vendor
            ws['E6'] = invoice_date_fmt
            ws['E7'] = invoice_number
            ws['B8'] = po_approved
            ws['B9'] = agreement_signed
            ws['E8'] = po_expiry_date
            ws['E9'] = agreement_signed_date
            ws['E11'] = msme
            ws['B13'] = isd  # Or any cell where ISD info should appear
            ws['E12'] = mobile_no
            ws['F25'] = total_amount
            ws['F25'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F25'].number_format = numbers.FORMAT_NUMBER_00

            ws['F35'] = total_amount
            ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F35'].number_format = numbers.FORMAT_NUMBER_00

            if ws['A35'].value:
                ws['A35'] = f"{ws['A35'].value} {total_amount_words}"
            else:
                ws['A35'] = total_amount_words
            ws['A35'].alignment = Alignment(wrap_text=True)
            ws.column_dimensions['A'].width = 50  # Increase the width of column A
            ws.row_dimensions[35].height = 40  # Increase the height of row 35

            if ws['A25'].value:  # Check if there is already a value in the cell
                ws['A25'] = f"{ws['A25'].value} ({tag1})"
            else:
                ws['A25'] = f"Marketing Expenses ({tag1})"

            if ws['D47'].value:  # Check if there's already a value in the cell
                ws['D47'] = f"{ws['D47'].value} {created_by}"
            else:
                ws['D47'] = created_by

            # For approved_by
            if ws['D41'].value:
                ws['D41'] = f"{ws['D41'].value} {approved_by}"
            else:
                ws['D41'] = approved_by

            # For reviewed_by
            if ws['D39'].value:
                ws['D39'] = f"{ws['D39'].value} {reviewed_by}"
            else:
                ws['D39'] = reviewed_by

            # For hod_approval
            if ws['D43'].value:
                ws['D43'] = f"{ws['D43'].value} {hod_approval}"
            else:
                ws['D43'] = hod_approval

            # For ceo_approval
            if ws['D45'].value:
                ws['D45'] = f"{ws['D45'].value} {ceo_approval}"
            else:
                ws['D45'] = ceo_approval
            
            # Save the populated Excel file to a BytesIO stream
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Return the Excel file as a response
            return send_file(
                excel_file,
                as_attachment=True,
                download_name=f"updated_invoice_{invoice_number}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    conn.close()

    return render_template(
        'edit_invoice.html', 
        invoice=invoice, 
        vendors=vendors, 
        approved_by_values=approved_by_values, 
        created_by_values=created_by_values,
        hod_values=hod_values,
        ceo_values=ceo_values,
        reviewed_by_values=reviewed_by_values,
        tag1=tag1,
        tag2=tag2
    )
    pass

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_invoice(id):
    # 🔐 Only Superadmin Can Delete
    if current_user.role != "superadmin":
        flash("Only Superadmin can delete invoices.")
        return redirect(url_for('index'))

    conn = get_db_connection()
    try:
        # 🔹 ALWAYS use a fresh cursor for DELETE
        cursor = conn.cursor(dictionary=True)

        # Optional: verify invoice exists
        cursor.execute("SELECT id, invoice_number FROM invoices WHERE id = %s AND deleted_at IS NULL", (id,))
        invoice = cursor.fetchone()

        if not invoice:
            flash("Invoice not found or already deleted.")
            return redirect(url_for('index'))

        # 🔥 IMPORTANT: consume results completely
        cursor.fetchall()

        cursor.execute("""
            UPDATE invoices
            SET deleted_at = NOW(),
                deleted_by = %s
            WHERE id = %s
        """, (current_user.email,id))

        conn.commit()

        log_activity(f"Soft-deleted invoice ID {id} ({invoice['invoice_number']})")
        flash("Invoice deleted successfully!")

    except Exception as e:
        conn.rollback()
        print("Delete operaion error:", {e})
        flash('Failed to delete invoice.')

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('index'))

@app.route('/manage_vendors', methods=['GET'])
@login_required
def manage_vendors():
    """Department landing or filtered vendor kanban depending on query param"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    user_role = current_user.role
    user_department = current_user.department

    selected_dept = request.args.get('department', '').strip()

    # Non-superadmin always goes straight to their own department
    if user_role != 'superadmin':
        selected_dept = user_department

    if selected_dept:
        # Vendor kanban view — filtered by department
        cursor.execute(
            'SELECT * FROM vendors WHERE deleted_at IS NULL AND department = %s ORDER BY vendor_name',
            (selected_dept,)
        )
        vendors = cursor.fetchall()
        conn.close()
        return render_template('manage_vendors.html',
                               vendors=vendors,
                               selected_department=selected_dept,
                               departments=None)
    else:
        # Department landing view — superadmin only
        # Fetch all departments from departments table + join vendor counts
        cursor.execute('''
            SELECT
                d.department_name AS department,
                COUNT(CASE WHEN v.deleted_at IS NULL THEN 1 END) AS total,
                SUM(CASE WHEN v.vendor_status = 'Active' AND v.deleted_at IS NULL THEN 1 ELSE 0 END) AS active_count,
                SUM(CASE WHEN v.vendor_status = 'Inactive' AND v.deleted_at IS NULL THEN 1 ELSE 0 END) AS inactive_count
            FROM departments d
            LEFT JOIN vendors v ON v.department = d.department_name
            GROUP BY d.id, d.department_name
            ORDER BY d.department_name
        ''')
        departments = cursor.fetchall()
        conn.close()
        return render_template('manage_vendors.html',
                               vendors=None,
                               selected_department=None,
                               departments=departments)

# ============== Department addition (superadmin only) ==================
@app.route('/vendor/add_department', methods=['POST'])
@superadmin_required
def add_department():
    """Create a new department in the departments table"""
    dept_name = request.form.get('dept_name', '').strip()
    if not dept_name:
        flash('Department name is required.', 'error')
        return redirect(url_for('manage_vendors'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Check if department already exists
    cursor.execute('SELECT id FROM departments WHERE department_name = %s', (dept_name,))
    existing = cursor.fetchone()

    if existing:
        conn.close()
        flash(f'Department "{dept_name}" already exists.', 'error')
        return redirect(url_for('manage_vendors'))

    # Insert into departments table
    cursor.execute('INSERT INTO departments (department_name) VALUES (%s)', (dept_name,))
    conn.commit()
    conn.close()

    flash(f'Department "{dept_name}" created successfully!', 'success')
    return redirect(url_for('manage_vendors'))

# ============= VENDOR BULK IMPORT ROUTES (superadmin only) =============
@app.route('/vendor/import/template')
@superadmin_required
def vendor_import_template():
    """Return a downloadable Excel template for vendor bulk import"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    headers = ['vendor_name', 'vendor_address', 'PAN', 'GSTIN', 'POC', 'POC_number', 'POC_email', 'description']
    ws.append(headers)
    # Sample row so user knows the format
    ws.append(['Example Vendor Pvt Ltd', '123 MG Road, Mumbai 400001', 'ABCDE1234F', '27ABCDE1234F1Z5', 'John Doe', '9876543210', 'john@example.com', 'Sample description'])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='vendor_import_template.xlsx'
    )


@app.route('/vendor/import/preview', methods=['POST'])
@superadmin_required
def vendor_import_preview():
    """Parse uploaded file, validate rows, return valid + error lists"""
    try:
        file = request.files.get('file')
        department = request.form.get('department', '').strip()

        if not file:
            return jsonify({'success': False, 'message': 'No file uploaded'})
        if not department:
            return jsonify({'success': False, 'message': 'Department is required'})

        filename = file.filename.lower()
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file)
        else:
            return jsonify({'success': False, 'message': 'Only .xlsx or .csv files are supported'})

        # Normalise column names
        df.columns = [c.strip().lower() for c in df.columns]

        if 'vendor_name' not in df.columns or 'vendor_address' not in df.columns:
            return jsonify({'success': False, 'message': 'File must have vendor_name and vendor_address columns'})

        # Fetch existing vendor names from DB to catch duplicates
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT LOWER(vendor_name) FROM vendors WHERE deleted_at IS NULL")
        existing = {row[0] for row in cursor.fetchall()}
        conn.close()

        valid, errors = [], []
        def clean(row, val):
            v = str(row.get(val, '')).strip()
            return None if (not v or v.lower() == 'nan') else v
        
        for i, row in df.iterrows():
            row_num = i + 2  # +2 because row 1 is header, pandas is 0-indexed
            vendor_name = str(row.get('vendor_name', '')).strip()
            vendor_address = str(row.get('vendor_address', '')).strip()

            # Validate required fields
            if not vendor_name or vendor_name.lower() == 'nan':
                errors.append({'row': row_num, 'vendor_name': '', 'vendor_address': vendor_address, 'reason': 'vendor_name is required'})
                continue
            if not vendor_address or vendor_address.lower() == 'nan':
                errors.append({'row': row_num, 'vendor_name': vendor_name, 'vendor_address': '', 'reason': 'vendor_address is required'})
                continue
            if vendor_name.lower() in existing:
                errors.append({'row': row_num, 'vendor_name': vendor_name, 'vendor_address': vendor_address, 'reason': 'Vendor already exists'})
                continue

            valid.append({
                'row': row_num,
                'vendor_name': vendor_name,
                'vendor_address': vendor_address,
                'department': department,   # always from the UI input
                'pan': clean(row,'pan'),
                'gstin': clean(row,'gstin'),
                'poc': clean(row,'poc'),
                'poc_number': clean(row,'poc_number'),
                'poc_email': clean(row,'poc_email'),
                'description': clean(row,'description'),
                'shortform': generate_shortform(vendor_name)
            })

        return jsonify({'success': True, 'valid': valid, 'errors': errors})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/vendor/import/confirm', methods=['POST'])
@superadmin_required
def vendor_import_confirm():
    """Bulk insert validated vendor rows into the vendors table"""
    try:
        data = request.get_json()
        rows = data.get('rows', [])

        if not rows:
            return jsonify({'success': False, 'message': 'No rows to import'})

        conn = get_db_connection()
        cursor = conn.cursor()

        imported = 0
        for row in rows:
            cursor.execute("""
                INSERT INTO vendors
                (vendor_name, vendor_status, department, description, shortforms_of_vendors,
                 vendor_address, PAN, GSTIN, POC, POC_number, POC_email)
                VALUES (%s, 'Active', %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                row['vendor_name'],
                row['department'],
                row.get('description'),
                row['shortform'],
                row['vendor_address'],
                row.get('pan'),
                row.get('gstin'),
                row.get('poc'),
                row.get('poc_number'),
                row.get('poc_email')
            ))
            imported += 1

        conn.commit()
        conn.close()

        log_activity(f"{current_user.name}({current_user.role}) bulk imported {imported} vendors into department '{rows[0]['department']}'")

        return jsonify({'success': True, 'imported': imported})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/edit_vendor/<int:id>', methods=['POST'])
@superadmin_required
def edit_vendor(id):
    """Edit existing vendor details"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get form data
    vendor_name = request.form.get('vendor_name')
    vendor_status = request.form.get('vendor_status')
    department = request.form.get('department')
    shortform = request.form.get('shortform')
    vendor_address = request.form.get('vendor_address')
    pan = request.form.get('pan')
    gstin = request.form.get('gstin')
    poc = request.form.get('poc')
    poc_number = request.form.get('poc_number')
    poc_email = request.form.get('poc_email')

    # Fetch old vendor name for logging
    cursor.execute('SELECT vendor_name FROM vendors WHERE id = %s', (id,))
    vendor = cursor.fetchone()
    old_vendor_name = vendor['vendor_name'] if vendor else 'Unknown Vendor'

    # Update vendor with all fields
    cursor.execute('''
        UPDATE vendors 
        SET vendor_name = %s,
            vendor_status = %s,
            department = %s,
            shortforms_of_vendors = %s,
            vendor_address = %s,
            PAN = %s,
            GSTIN = %s,
            POC = %s,
            POC_number = %s,
            POC_email = %s
        WHERE id = %s
    ''', (vendor_name, vendor_status, department, shortform, vendor_address, 
          pan, gstin, poc, poc_number, poc_email, id))
    
    conn.commit()
    conn.close()
    
    role = current_user.role
    user_name = current_user.name
    if old_vendor_name == vendor_name:
        log_activity(f"{user_name}({role}) updated vendor '{old_vendor_name}' to '{vendor_name}'")
    else:
        log_activity(f"{user_name}({role}) updated vendor '{vendor_name}'")
    flash('Vendor updated successfully!')
    return redirect(url_for('manage_vendors'))

# ============= VENDOR APPROVAL ROUTES =============

@app.route('/vendor/request', methods=['POST'])
@login_required
def request_vendor():
    """User submits vendor request for approval"""
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO vendor_requests 
            (vendor_name, description, department, vendor_address, PAN, GSTIN, 
             POC, POC_number, POC_email, requested_by_user_id, requested_by_name, 
             requested_by_email, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending')
        """, (
            data['vendor_name'],
            data.get('description'),
            data.get('department'),
            data.get('vendor_address'),
            data.get('pan'),
            data.get('gstin'),
            data.get('poc'),
            data.get('poc_number'),
            data.get('poc_email'),
            current_user.id,
            current_user.name,
            current_user.email
        ))
        
        conn.commit()
        conn.close()
        
        log_activity(f"{current_user.name} requested new vendor: {data['vendor_name']}")
        
        return jsonify({
            'success': True,
            'message': 'Vendor request submitted! Waiting for admin approval.'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= VENDOR APPROVAL HELPER FUNCTION =============

def generate_shortform(vendor_name):
    """Generate shortform from vendor name - first letter of each word"""
    words = vendor_name.strip().split()
    shortform = ''.join([word[0].upper() for word in words if word])
    return shortform[:10]  # Max 10 characters

@superadmin_required
@app.route('/approvals')
@login_required
def approvals():
    """Display all pending approvals (superadmin only)"""
    if current_user.role != 'superadmin':
        flash("You don't have permission to access this page")
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT * FROM vendor_requests
        WHERE status = 'pending'
        ORDER BY request_date DESC
    """)
    
    requests = cursor.fetchall()
    conn.close()
    
    return render_template('approvals.html', requests=requests)


@app.route('/api/pending-count')
@login_required
def pending_count():
    """Get count of pending vendor requests"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM vendor_requests WHERE status = 'pending'")
    count = cursor.fetchone()[0]
    conn.close()
    
    return jsonify({'count': count})


@app.route('/vendor/request-details/<int:request_id>')
@login_required
def vendor_request_details(request_id):
    """Get vendor request details"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM vendor_requests WHERE id = %s", (request_id,))
    details = cursor.fetchone()
    conn.close()
    
    if details:
        details['request_date'] = details['request_date'].isoformat()
        return jsonify(details)
    return jsonify({'error': 'Not found'}), 404

def extract_name_from_email(email):
    """Extract name from email: tushar.kadam@auxilo.com -> Tushar Kadam"""
    try:
        username = email.split('@')[0]
        name_parts = username.split('.')
        formatted_name = ' '.join(part.capitalize() for part in name_parts)
        return formatted_name
    except:
        return "Team"
    
@app.route('/vendor/approve/<int:request_id>', methods=['POST'])
@login_required
def approve_vendor_request(request_id):
    """Superadmin approves vendor request"""
    if current_user.role != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM vendor_requests WHERE id = %s", (request_id,))
        vendor_req = cursor.fetchone()
        
        if not vendor_req:
            return jsonify({'success': False, 'message': 'Request not found'}), 404
        
        # Insert into vendors table
        cursor.execute("""
            INSERT INTO vendors 
            (vendor_name, vendor_status, department, description, shortforms_of_vendors,
            vendor_address, PAN, GSTIN, POC, POC_number, POC_email)
            VALUES (%s, 'Active', %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            vendor_req['vendor_name'],
            vendor_req['department'],
            vendor_req['description'],  # Full description
            generate_shortform(vendor_req['vendor_name']),  # Generate shortform from vendor name
            vendor_req['vendor_address'],
            vendor_req['PAN'],
            vendor_req['GSTIN'],
            vendor_req['POC'],
            vendor_req['POC_number'],
            vendor_req['POC_email']
        ))
        
        # Update request status
        cursor.execute("""
            UPDATE vendor_requests 
            SET status = 'approved',
                reviewed_by_user_id = %s,
                reviewed_by_name = %s,
                reviewed_date = NOW()
            WHERE id = %s
        """, (current_user.id, current_user.name, request_id))
        
        conn.commit()
        conn.close()
        
        log_activity(f"{current_user.name} approved vendor: {vendor_req['vendor_name']}")

        email_sent = False
        try:
            recipients = os.getenv('VENDOR_ADDED_RECIPIENTS', '').split(',')
            for recipient in [r.strip() for r in recipients if r.strip()]:
                sent = email_service.send_vendor_approved(
                    email=recipient,
                    vendor_name=vendor_req['vendor_name'],
                    description=vendor_req['description'],
                    requested_by=vendor_req['requested_by_name'],
                    approved_by=current_user.name
                )
                if sent:
                    email_sent=True
                    logger.info(f"Vendor approved email sent to {recipient}")
                else:
                    logger.warning(f"Vendor approved email FAILED for {recipient}")
                     
            email_sent = True
            log_activity(f'Email sent: vendor {vendor_req["vendor_name"]} approved')
        except Exception as e:
            logger.error(f"Failed to send vendor approval email: {e}")
            email_sent = False
        return jsonify({
            'success': True,
            'message': f"Vendor '{vendor_req['vendor_name']}' approved successfully!" + 
                    ("" if email_sent else " (Email notification failed)"),
            'email_sent': email_sent
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/vendor/reject/<int:request_id>', methods=['POST'])
@login_required
def reject_vendor_request(request_id):
    """Superadmin rejects vendor request"""
    if current_user.role != 'superadmin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        reason = data.get('reason', 'No reason provided')
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM vendor_requests WHERE id = %s", (request_id,))
        vendor_req = cursor.fetchone()
        
        if not vendor_req:
            return jsonify({'success': False, 'message': 'Request not found'}), 404
        
        cursor.execute("""
            UPDATE vendor_requests 
            SET status = 'rejected',
                reviewed_by_user_id = %s,
                reviewed_by_name = %s,
                reviewed_date = NOW(),
                rejection_reason = %s
            WHERE id = %s
        """, (current_user.id, current_user.name, reason, request_id))
        
        conn.commit()
        conn.close()
        
        log_activity(f"{current_user.name} rejected vendor: {vendor_req['vendor_name']}")
        
        return jsonify({
            'success': True,
            'message': 'Vendor request rejected'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/vendor/delete/<int:id>', methods=['POST'])
@login_required
def soft_delete_vendor(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Soft delete - set deleted_at to current timestamp
        cursor.execute("""
            UPDATE vendors 
            SET deleted_at = NOW() 
            WHERE id = %s
        """, (id,))
        
        conn.commit()
        conn.close()
        
        # Log activity
        user = current_user.name
        role = current_user.role
        log_activity(f"{user}({role}) soft deleted vendor ID {id}")
        
        flash("Vendor deleted successfully", "success")
        return redirect(url_for('manage_vendors'))
        
    except Exception as e:
        flash(f"Error deleting vendor: {str(e)}", "danger")
        return redirect(url_for('manage_vendors'))  
# Route to manage dropdowns (add new values)
@app.route('/manage_dropdowns', methods=['GET', 'POST'])
@superadmin_required
def manage_dropdowns():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    user_name = current_user.name
    role = current_user.role

    if request.method == 'POST':
        type_ = request.form.get('type')
        value = request.form.get('value')
        if type_ and value:
            cursor.execute(
                'INSERT INTO dropdown_values (type, value) VALUES (%s, %s)', 
                (type_, value)
            )
            conn.commit()
            flash('Value added successfully.')

            #  Log the activity
            actor_role = current_user.role
            user_name = request.cookies.get('name') or request.cookies.get('email')
            log_activity(f"{user_name}({actor_role}) added a new value '{value}' to the '{type_}' dropdown")

    # Load all dropdowns
    cursor.execute('SELECT DISTINCT type FROM dropdown_values')
    types = [row['type'] for row in cursor.fetchall()]

    dropdown_data = {}
    for t in types:
        cursor.execute(
            'SELECT id, value FROM dropdown_values WHERE type = %s AND is_active = TRUE', 
            (t,)
        )
        dropdown_data[t] = cursor.fetchall()

    conn.close()
    return render_template('manage_dropdowns.html', dropdown_data=dropdown_data)


#Delete dropdown value from Database(soft delete)
@app.route('/delete_dropdown/<int:id>', methods=['POST'])
@superadmin_required
def delete_dropdown(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch the value before updating for logging
    cursor.execute('SELECT type, value FROM dropdown_values WHERE id = %s', (id,))
    value_data = cursor.fetchone()

    if value_data:
        cursor.execute('UPDATE dropdown_values SET is_active = FALSE WHERE id = %s', (id,))
        conn.commit()

        #  Log the activity
        role = current_user.role
        user_name = request.cookies.get('name') or request.cookies.get('email')
        log_activity(f"{user_name}({role}) soft-deleted the value '{value_data['value']}' from '{value_data['type']}' dropdown")

        flash('Value disabled (soft-deleted).')
    else:
        flash('Dropdown value not found.')

    conn.close()
    return redirect(url_for('manage_dropdowns'))


# Route to manage users (superadmin only)
@app.route('/manage_users')
@superadmin_required
def manage_users():
    users = Users.query.all()
    from sqlalchemy import text
    departments = db.session.execute(text('SELECT department_name FROM departments ORDER BY department_name')).fetchall()
    departments = [d[0] for d in departments]
    user_name = request.cookies.get('name') or request.cookies.get('email')
    return render_template('manage_users.html',users=users,departments=departments)

# Route to add a new user (superadmin only)
@app.route('/add_user', methods=['POST'])
@superadmin_required
def add_user():
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    role = request.form.get('role', '').strip()
    department = request.form.get('department', '').strip()

    if not name or len(name) <2:
        flash('Name must be at least 2 characters long','error')
        return redirect(url_for('manage_users'))
    
    if not email or '@' not in email:
        flash('Invalid email address','error')
        return redirect(url_for('manage_users'))
    
    if role not in ['user', 'admin', 'superadmin']:
        flash('Invalid role','error')
        return redirect(url_for('manage_users'))

    existing_user = Users.query.filter_by(email=email).first()
    if not existing_user:
        new_user = Users(name=name, email=email, role=role,department=department, is_active=True)
        db.session.add(new_user)
        db.session.commit()

        #  Log activity
        actor_role = current_user.role
        actor = request.cookies.get('name') or request.cookies.get('email')
        log_activity(f"{actor}({actor_role}) added new user: {name} ({email}) with role {role}")

    return redirect(url_for('manage_users'))

# Route to edit/delete user details (superadmin only)
@app.route('/delete_user/<int:user_id>', methods=['POST'])
@superadmin_required
def delete_user(user_id):
    user = Users.query.get_or_404(user_id)

    #  Get actor BEFORE deletion
    actor = request.cookies.get('name') or request.cookies.get('email')

    #  Log BEFORE deletion
    role = current_user.role
    log_activity(f"{actor}({role}) deleted user: {user.name} ({user.email})")

    db.session.delete(user)
    db.session.commit()

    return redirect(url_for('manage_users'))

# Route to toggle user status (active/inactive) (superadmin only)
@app.route('/toggle_user_status/<int:user_id>', methods=['POST'])
@superadmin_required
def toggle_user_status(user_id):
    user = Users.query.get_or_404(user_id)

    # Toggle the status
    user.is_active = not user.is_active
    db.session.commit()

    # Identify actor from cookies
    actor = request.cookies.get('name') or request.cookies.get('email')

    # Log the action
    status = "activated" if user.is_active else "deactivated"
    log_activity(f"{actor} {status} user: {user.name} ({user.email})")

    return redirect(url_for('manage_users'))

# Route to update user role (superadmin only)
@app.route('/update_user_role/<int:user_id>', methods=['POST'])
@superadmin_required
def update_user_role(user_id):
    user = Users.query.get_or_404(user_id)
    new_role = request.form.get('role')
    new_name = request.form.get('name', '').strip()
    new_email = request.form.get('email','').strip()
    new_status = request.form.get('status')
    new_department = request.form.get('department','').strip()

    if new_role in ['user', 'admin', 'superadmin']:
        old_role = user.role
        user.role = new_role

        if new_name:
            user.name = new_name

        if new_email:
            user.email = new_email
            
        if new_department:
            user.department = new_department

        if new_status in ['active','inactive']:
            user.is_active = (new_status == 'active')

        db.session.commit()

        # Identify actor from cookie
        actor = request.cookies.get('name') or request.cookies.get('email')

        # Log the role change
        actor_role = current_user.role
        log_activity(f"{actor}({actor_role}) changed role of user {user.name} ({user.email}) from {old_role} to {new_role}")

    return redirect(url_for('manage_users'))

# Route to view activity logs (superadmin only)
@app.route('/activity_logs')
@superadmin_required
def view_activity_logs():
    # Only show logs from the last 15 days
    fifteen_days_ago = datetime.now() - timedelta(days=15)
    logs = ActivityLog.query.filter(ActivityLog.timestamp >= fifteen_days_ago).order_by(ActivityLog.timestamp.desc()).all()
    return render_template('activity_logs.html', logs=logs)

# Route to download activity logs as CSV (superadmin only)
@app.route('/download_activity_logs')
@superadmin_required
def download_activity_logs():
    """
    Download activity logs with filtering options:
    - All time 
    - Financial Year
    - Specific Month(s)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # All activity logs
    filter_type = request.args.get('filter_type','all')
    # Financial year activity logs
    fy_year = request.args.get('fy_year','')
    # Month filtered activity logs
    month_start = request.args.get('month_start','')
    month_end = request.args.get('month_end','')

    query = ActivityLog.query

    filter_description = "All Time"

    if filter_type == 'fy' and fy_year:
        try:
            start_year = int(fy_year.split('-')[0])
            fy_start = datetime(start_year,4,1)
            fy_end = datetime(start_year + 1,3,31,23,59,59)

            query = query.filter(
                ActivityLog.timestamp >= fy_start,
                ActivityLog.timestamp <= fy_end
            )
            filter_description = f"Financial Year {fy_year}"
        
        except (ValueError, IndexError):
            flash('Invalid financial year format', 'error')
            return redirect(url_for('view_activity_logs'))

    elif filter_type == 'month' and month_start:
        try:
            start_year,start_month = map(int, month_start.split('-'))
            range_start = datetime(start_year, start_month, 1)

            if month_end:
                end_year, end_month = map(int, month_end.split('-'))
            else:
                end_year, end_month = start_year, start_month

            if end_month == 12:
                range_end = datetime(end_year + 1, 1, 1) - timedelta(days=1)
            else:
                range_end = datetime(end_year, end_month + 1, 1) - timedelta(days=1)

            range_end = range_end.replace(hour=23,minute=59,second=59)

            query = query.filter(
                ActivityLog.timestamp >= range_start,
                ActivityLog.timestamp <= range_end
            )

            if month_start == month_end or not month_end:
                filter_description = range_start.strftime('%B %Y')
            else:
                filter_description = f"{range_start.strftime('%B %Y')} to {range_end.strftime('%B %Y')}"

        except (ValueError, IndexError):
            flash('Invalid month format', 'error')
            return redirect(url_for('view_activity_logs')) 
          
    # Gets filtered logs
    logs = query.order_by(ActivityLog.timestamp.desc()).all()

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Logs"

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # TITLE AND METADATA
    # ============================================
    
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "ACTIVITY LOGS REPORT"
    title_cell.font = Font(bold=True, size=16, color="366092")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:F2')
    filter_cell = ws['A2']
    filter_cell.value = f"Filter: {filter_description}"
    filter_cell.font = Font(italic=True, size=11)
    filter_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:F3')
    date_cell = ws['A3']
    date_cell.value = f"Generated on: {datetime.now().strftime('%d %B %Y at %I:%M %p')}"
    date_cell.font = Font(italic=True, size=10)
    date_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A4:F4')
    count_cell = ws['A4']
    count_cell.value = f"Total Records: {len(logs)}"
    count_cell.font = Font(bold=True, size=11, color="0066CC")
    count_cell.alignment = Alignment(horizontal="center")
    
    ws.row_dimensions[5].height = 5

    # ============================================
    # HEADERS
    # ============================================
    
    headers = ['ID', 'User Email', 'User Role', 'Action', 'Department', 'Timestamp']
    ws.append(headers)
    
    header_row = ws[6]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ============================================
    # DATA ROWS
    # ============================================
    
    for log in logs:
        # Extract role from action if available (format: "Name(role) did something")
        import re
        role_match = re.search(r'\(([^)]+)\)', log.action)
        user_role = role_match.group(1) if role_match else 'N/A'
        
        ws.append([
            log.id,
            log.user_email,
            user_role,
            log.action,
            log.department if hasattr(log, 'department') else 'N/A',
            log.timestamp.strftime('%d-%m-%Y %H:%M:%S')
        ])
    
    # ============================================
    # FORMAT DATA ROWS
    # ============================================
    
    # Alignment for data rows
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    timestamp_alignment = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=6):
        for idx, cell in enumerate(row):
            cell.border = thin_border
            
            # Center align ID and Timestamp
            if idx in [0, 5]:
                cell.alignment = timestamp_alignment
            else:
                cell.alignment = data_alignment
            
            # Zebra striping
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # ============================================
    # COLUMN WIDTHS
    # ============================================
    
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 30  # Email
    ws.column_dimensions['C'].width = 15  # Role
    ws.column_dimensions['D'].width = 60  # Action
    ws.column_dimensions['E'].width = 15  # Department
    ws.column_dimensions['F'].width = 20  # Timestamp
    
    # Set row heights
    ws.row_dimensions[1].height = 25  # Title
    ws.row_dimensions[6].height = 30  # Headers

    # ============================================
    # SAVE AND SEND
    # ============================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename based on filter
    if filter_type == 'fy':
        filename = f"activity_logs_FY{fy_year}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    elif filter_type == 'month':
        filename = f"activity_logs_{month_start}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    else:
        filename = f"activity_logs_all_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Route for user login
@app.route('/logout')
@login_required
def logout():
    try:
        user = current_user.name
        role = current_user.role
    except:
        user = "Unknown"
        role = "Unknown"

    # Log activity BEFORE logout
    log_activity(f"{user}({role}) logged out")

    # Securely log user out
    logout_user()

    flash("Logged out successfully.")

    return redirect(url_for('login'))
# Middleware to add headers to prevent caching
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# @app.route('/test_monthly_report')
# @login_required
# def test_monthly_report():
#     run_monthly_report()
#     return "Monthly report sent successfully (TEST MODE)"

def run_monthly_report():
    with app.app_context():
        summary, invoices, start_date, end_date = get_monthly_summary()
        excel = create_monthly_excel(summary, invoices, start_date, end_date)
        send_monthly_email(excel, start_date)

        # Optional log
        log_activity(f"System sent monthly report for {start_date.strftime('%B %Y')}")

@app.route('/download_single_excel/<int:id>', methods=['GET'])
@login_required
def download_single_excel(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM invoices WHERE id = %s", (id,))
    invoice = cursor.fetchone()

    if not invoice:
        conn.close()
        return "Invoice not found", 404

    conn.close()

    import pandas as pd
    from io import BytesIO

    df = pd.DataFrame([invoice])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=f"invoice_{invoice['invoice_number']}.xlsx",
        as_attachment=True
    )

# ==================== PO SYSTEM ROUTES ====================
# Add these routes to your app_duplicate.py file

@app.route('/po/list', methods=['GET'])
@login_required
def po_list():
    """Display all Purchase Orders"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Fetch all POs with vendor details
    cursor.execute("""
    SELECT 
        po.id,
        po.po_number,
        po.po_date,
        po.grand_total,
        po.created_at,
        v.vendor_name,
        u_approved.name as approved_by_name,
        u_reviewed.name as reviewed_by_name
    FROM invoice_uat_db.purchase_orders po
    LEFT JOIN invoice_uat_db.vendors v ON po.vendor_id = v.id
    LEFT JOIN invoice_uat_db.users u_approved ON po.approved_by = u_approved.id
    LEFT JOIN invoice_uat_db.users u_reviewed ON po.reviewed_by = u_reviewed.id
    ORDER BY po.created_at DESC
""")

    pos = cursor.fetchall()

    cursor.execute("""
        SELECT id, vendor_name, vendor_address, PAN, GSTIN, 
               POC, POC_number, POC_email,department 
        FROM invoice_uat_db.vendors
    """)
    vendors = cursor.fetchall()
    
    # Format dates
    for po in pos:
        if po['po_date']:
            po['po_date'] = po['po_date'].strftime('%d/%m/%Y')
        if po['created_at']:
            po['created_at'] = po['created_at'].strftime('%d/%m/%Y %H:%M:%S')
    
    conn.close()
    return render_template('po_list.html', pos=pos, vendors=vendors)


@app.route('/po/add', methods=['POST'])
@login_required
def add_po():
    """Create a new Purchase Order"""
    try:
        data = request.get_json()

        # Logged in user
        user_email = current_user.email
        user_name = current_user.name
        role = current_user.role

        if not user_email:
            return jsonify({'success': False, 'message': 'User not authenticated'}), 401

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # ✅ Handle optional PO number and date
        po_number = data.get('po_number')
        po_date_raw = data.get('po_date')
        
        # Check duplicate PO number only if provided
        if po_number:
            cursor.execute(
                'SELECT id FROM invoice_uat_db.purchase_orders WHERE po_number = %s',
                (po_number,)
            )
            if cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': f"PO number {po_number} already exists"}), 400

        # ✅ Convert date DD/MM/YYYY → YYYY-MM-DD (only if provided)
        if po_date_raw:
            po_date_parts = po_date_raw.split('/')
            po_date_db = f"{po_date_parts[2]}-{po_date_parts[1]}-{po_date_parts[0]}"
        else:
            po_date_db = None

        # Vendor handling
        cursor.execute(
            'SELECT id FROM invoice_uat_db.vendors WHERE vendor_name = %s',
            (data['vendor_name'],)
        )
        vendor = cursor.fetchone()

        if not vendor:
            cursor.execute(
                'INSERT INTO invoice_uat_db.vendors (vendor_name, vendor_address) VALUES (%s, %s)',
                (data['vendor_name'], data['vendor_address'])
            )
            vendor_id = cursor.lastrowid
        else:
            vendor_id = vendor['id']

        # User_PO mapping
        cursor.execute(
            'SELECT id FROM invoice_uat_db.users WHERE email = %s',
            (user_email,)
        )
        user_po = cursor.fetchone()

        if not user_po:
            cursor.execute(
                'INSERT INTO invoice_uat_db.users (email, name, role) VALUES (%s, %s, %s)',
                (user_email, user_name or user_email.split("@")[0], role or "user")
            )
            created_by_id = cursor.lastrowid
        else:
            created_by_id = user_po['id']

        # Default approver
        default_email = 'abhilash.pillai@auxilo.com'
        cursor.execute(
            'SELECT id FROM invoice_uat_db.users WHERE email = %s',
            (default_email,)
        )
        default_user = cursor.fetchone()

        if not default_user:
            cursor.execute(
                'INSERT INTO invoice_uat_db.users (email, name, role) VALUES (%s, %s, %s)',
                (default_email, 'Abhilash Pillai', 'approver')
            )
            default_user_id = cursor.lastrowid
        else:
            default_user_id = default_user['id']

        # Totals
        total_amount = total_cgst = total_sgst = grand_total = 0

        for item in data['items']:
            total = float(item['total'])
            base = total / 1.18
            cgst = base * 0.09
            sgst = base * 0.09

            total_amount += base
            total_cgst += cgst
            total_sgst += sgst
            grand_total += total

        amount_words = amount_to_words(grand_total)

        # ✅ Do NOT auto-generate PO number - keep it NULL if not provided
        # The frontend explicitly sends null when "PO is not mandatory"
        
        # ✅ Generate PDF path - use temp name if no PO number
        if po_number:
            pdf_filename = f"{po_number.replace('/', '_')}.pdf"
        else:
            # Use timestamp for POs without numbers
            from datetime import datetime
            pdf_filename = f"PO_NoNumber_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        pdf_path = f"generated_pdfs/{pdf_filename}"
        
        # Insert PO
        try:
            # Start transaction
            conn.autocommit = False
            
            cursor.execute("""
                INSERT INTO invoice_uat_db.purchase_orders (
                    po_number, vendor_id, po_date, total_amount, cgst_amount, sgst_amount,
                    grand_total, pdf_path, approved_by, reviewed_by, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                po_number, vendor_id, po_date_db, total_amount, total_cgst,
                total_sgst, grand_total,
                pdf_path,
                default_user_id, default_user_id, created_by_id
            ))

            po_id = cursor.lastrowid

            # Insert PO items
            for item in data['items']:
                total = float(item['total'])
                base = total / 1.18
                qty = float(item.get('qty', 0)) if item.get('qty') else 0
                rate = float(item.get('rate', 0)) if item.get('rate') else (base / qty if qty else 0)

                cursor.execute("""
                    INSERT INTO invoice_uat_db.purchase_order_items (
                        po_id, product_description, quantity, rate, line_total
                    ) VALUES (%s, %s, %s, %s, %s)
                """, (po_id, item['description'], qty, rate, total))

            # Commit transaction
            conn.commit()
            conn.autocommit = True
            
        except Exception as e:
            # Rollback on error
            conn.rollback()
            conn.autocommit = True
            conn.close()
            logger.error(f"PO creation failed: {e}")
            return jsonify({'success': False, 'message': f'Transaction failed: {str(e)}'}), 500

        # PDF Data
        pdf_data = {
            "po_number": po_number or "N/A",  # Show N/A if no PO number
            "date": po_date_raw or date.today().strftime('%d/%m/%Y'),
            "vendor_address": data['vendor_address'],
            "items": [],
            "grand_total": grand_total,
            "amount_words": amount_words
        }

        for item in data['items']:
            total = float(item['total'])
            base = total / 1.18
            cgst = base * 0.09
            sgst = base * 0.09
            qty = float(item.get('qty', 0)) if item.get('qty') else 0
            rate = float(item.get('rate', 0)) if item.get('rate') else (base / qty if qty else 0)

            pdf_data["items"].append({
                "description": item["description"],
                "qty": qty,
                "rate": rate,
                "cgst": cgst,
                "sgst": sgst,
                "total": total
            })

        pdf_path = generate_po_pdf_flask(pdf_data)
        # Update DB with the actual pdf_path that was generated
        conn2 = get_db_connection()
        cursor2 = conn2.cursor()
        cursor2.execute(
            "UPDATE invoice_uat_db.purchase_orders SET pdf_path = %s WHERE id = %s",
            (pdf_path, po_id)
        )
        conn2.commit()
        conn2.close()

        conn.commit()
        conn.close()

        # Activity tracking
        if user_name and role:
            log_msg = f"{user_name}({role}) created PO"
            if po_number:
                log_msg += f" {po_number}"
            log_activity(log_msg)

        return jsonify({
            'success': True,
            'message': 'Purchase Order created successfully!',
            'po_id': po_id,
            'po_number': po_number or 'No PO Number'
        })

    except Exception as e:
        print("Error creating PO:", e)
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/po/download/<int:po_id>')
@login_required
def download_po_pdf(po_id):
    """Download PO PDF"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        'SELECT po_number, pdf_path FROM invoice_uat_db.purchase_orders WHERE id = %s',
        (po_id,)
    )
    po = cursor.fetchone()
    conn.close()

    if not po:
        flash('Purchase Order not found')
        return redirect(url_for('po_list'))

    pdf_path = po.get('pdf_path')

    # If path not stored, try building from po_number
    if not pdf_path and po.get('po_number'):
        safe_name = po['po_number'].replace('/', '_')
        pdf_path = f"generated_pdfs/{safe_name}.pdf"

    if not pdf_path or not os.path.exists(pdf_path):
        flash('PDF file not found')
        return redirect(url_for('po_list'))

    # Build download filename safely
    if po.get('po_number'):
        download_name = f"{po['po_number'].replace('/', '_')}.pdf"
    else:
        download_name = os.path.basename(pdf_path)

    return send_file(
        pdf_path,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=download_name
    )

@app.route('/po/detail/<int:po_id>')
@login_required
def po_detail(po_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
       SELECT 
            po.id,
            po.po_number,
            po.po_date,
            po.grand_total,
            po.vendor_id,
            v.vendor_name,
            v.vendor_address
        FROM purchase_orders po
        LEFT JOIN vendors v ON po.vendor_id = v.id
        WHERE po.id = %s
    """, (po_id,))
    
    po = cursor.fetchone()
    if not po:
        conn.close()
        return jsonify({"error": "PO not found"}), 404

    po["po_date"] = po["po_date"].strftime("%Y-%m-%d") if po["po_date"] else None

    cursor.execute("""
        SELECT 
            id,
            product_description,
            quantity,
            rate,
            line_total
        FROM purchase_order_items
        WHERE po_id = %s
    """, (po_id,))
    
    po["items"] = cursor.fetchall()

    conn.close()
    return jsonify(po)


@app.route('/po/update/<int:po_id>', methods=['POST'])
@login_required
def update_po(po_id):
    """Update existing Purchase Order"""
    try:
        data = request.get_json()

        # Logged in user
        user_email = request.cookies.get('email')
        user_name = request.cookies.get('name')
        role = request.cookies.get('role')

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Get PO number
        cursor.execute(
            'SELECT po_number FROM invoice_uat_db.purchase_orders WHERE id = %s',
            (po_id,)
        )
        po = cursor.fetchone()

        if not po:
            conn.close()
            return jsonify({'success': False, 'message': 'PO not found'}), 404

        po_number = po['po_number']

        # Delete existing items
        cursor.execute(
            'DELETE FROM invoice_uat_db.purchase_order_items WHERE po_id = %s',
            (po_id,)
        )

        # Recalculate totals
        total_amount = total_cgst = total_sgst = grand_total = 0

        for item in data['items']:
            total = float(item['total'])
            base = total / 1.18
            cgst = base * 0.09
            sgst = base * 0.09

            total_amount += base
            total_cgst += cgst
            total_sgst += sgst
            grand_total += total

            # Insert new items
            qty = float(item.get('qty', 0)) if item.get('qty') else 0
            rate = (
                float(item.get('rate', 0))
                if item.get('rate')
                else (base / qty if qty else 0)
            )

            cursor.execute("""
                INSERT INTO invoice_uat_db.purchase_order_items 
                (po_id, product_description, quantity, rate, line_total)
                VALUES (%s, %s, %s, %s, %s)
            """, (po_id, item['description'], qty, rate, total))

        # Update PO totals
        cursor.execute("""
            UPDATE invoice_uat_db.purchase_orders
            SET total_amount = %s,
                cgst_amount = %s,
                sgst_amount = %s,
                grand_total = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (total_amount, total_cgst, total_sgst, grand_total, po_id))

        conn.commit()
        conn.close()

        # ===============================
        #  CENTRAL ACTIVITY LOG
        # ===============================
        if user_name and role:
            log_activity(f"{user_name}({role}) updated PO {po_number}")

        return jsonify({
            'success': True,
            'message': 'Purchase Order updated successfully!'
        })

    except Exception as e:
        print(f"Error updating PO: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/po/activities')
@login_required
def po_activities():
    """Display PO activity logs"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get activities from last 30 days
    cursor.execute("""
        SELECT id, user_email, po_number, action, action_timestamp
        FROM invoice_uat_db.activity_of_po
        WHERE action_timestamp >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        ORDER BY action_timestamp DESC
        LIMIT 100
    """)
    
    activities = cursor.fetchall()
    
    for activity in activities:
        if activity['action_timestamp']:
            activity['action_timestamp'] = activity['action_timestamp'].strftime('%d/%m/%Y %H:%M:%S')
    
    conn.close()
    
    return render_template('po_activities.html', activities=activities)

@app.route('/po/delete/<int:po_id>', methods=['POST'])
@login_required
def delete_po(po_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get PO number before delete
        cursor.execute("SELECT po_number FROM invoice_uat_db.purchase_orders WHERE id=%s",
                       (po_id,))
        po = cursor.fetchone()

        # Delete items and PO
        cursor.execute("DELETE FROM invoice_uat_db.purchase_order_items WHERE po_id=%s",(po_id,))
        cursor.execute("DELETE FROM invoice_uat_db.purchase_orders WHERE id=%s",(po_id,))
        conn.commit()
        conn.close()

        user = current_user.name
        role = current_user.role
        log_activity(f"{user}({role}) deleted PO {po[0]}")

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

# ============================================================================
# GLOBAL ERROR HANDLERS
# ============================================================================

@app.errorhandler(400)
def bad_request(e):
    """Handle bad request errors"""
    return jsonify({
        'error': 'Invalid request. Please check your input and try again.',
        'details': str(e) if app.debug else None
    }), 400

@app.errorhandler(403)
def forbidden(e):
    """Handle forbidden errors"""
    flash("You don't have permission to access this resource.")
    return redirect(url_for('index'))

@app.errorhandler(404)
def not_found(e):
    """Handle not found errors"""
    flash("The requested page was not found.")
    return redirect(url_for('index'))

@app.errorhandler(500)
def internal_error(e):
    """Handle internal server errors"""
    logger.error(f"Internal error: {str(e)}")
    flash("An unexpected error occurred. Please try again or contact support.")
    return redirect(url_for('index'))

@app.errorhandler(Exception)
def handle_exception(e):
    """Handle all other exceptions"""
    logger.error(f"Unhandled exception: {str(e)}", exc_info=True)
    
    # Don't expose internal errors to users
    if app.debug:
        raise e
    else:
        flash("An error occurred. Please try again.")
        return redirect(url_for('index'))
    
@app.route('/api/total_logs_count')
@login_required
def get_total_logs_count():
    """Get total count of activity logs"""
    try:
        total = ActivityLog.query.count()
        return jsonify({'count': total})
    except Exception as e:
        logger.error(f"Error getting logs count: {e}")
        return jsonify({'count': 0})
    
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
